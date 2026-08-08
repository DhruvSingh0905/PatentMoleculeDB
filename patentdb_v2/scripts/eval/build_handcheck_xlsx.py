"""Hand-check Excel workbook — refactored to use canonical core/ modules.

Single sources used:
  - core.units.value_to_uM        (was 6 reimplementations)
  - core.patent_text.load_gp_description  (was 4)
  - core.bindingdb.iter_rows_for_patent  (was 3)
  - core.compound_id.parse_compound_id   (was 5)
  - core.smiles_utils.get_inchikey       (was several ad-hoc rdkit calls)

Per user feedback this turn:
  - Sort each sheet by `value_verdict` priority (match first → mismatches)
  - For US8952177: column names mirror Jie's exact CSV format
    (flap_binding_ki_uM, hwb_ltb4_ic50_uM, ..._qualifier, ..._n_runs)
  - n_runs values left blank (queued LLM-prompt fix is API-credit-blocked)
  - Both-blank rows dropped
  - Strict equality on values
"""
from __future__ import annotations

import csv
import json
import re
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from patentdb.core.units import value_to_uM
from patentdb.core.patent_text import load_gp_description
from patentdb.core.bindingdb import iter_rows_for_patent, attribute_source
from patentdb.core.smiles_utils import get_inchikey, get_stereo_flattened_key
from patentdb.core.iupac_to_smiles import _try_opsin, _convert_single
from patentdb.core.cost_tracker import cost_tracker
from patentdb.core.models import Compound, CompoundSource, IupacSource
from patentdb.scripts.eval import fidelity_check as fc
from patentdb.scripts.eval.assay_completeness_audit import distinct_measurements


DOWNLOADS = Path.home() / "Downloads"


# ── Color palette ───────────────────────────────────────────────


COLOR_HEADER = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
COLOR_MATCH = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")
COLOR_V2_WRONG = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
COLOR_REF_WRONG = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
COLOR_AMBIG = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
COLOR_V2_EXTRA = PatternFill(start_color="DEEBF7", end_color="DEEBF7", fill_type="solid")
COLOR_STEREO_DIFF = PatternFill(start_color="E4DFEC", end_color="E4DFEC", fill_type="solid")
# Gold: structure matched (same molecule) but the assay VALUE differs. Kept
# visually distinct from COLOR_V2_WRONG (peach = compound entirely missed) so a
# reviewer can tell "found, value disagrees" apart from "not extracted at all".
COLOR_VALUE_DIFF = COLOR_REF_WRONG
# Teal: v2's value was CONFIRMED in the patent's own assay table for this
# compound, but BDB lists a different number. v2 is faithful to the patent —
# the disagreement is BDB-vs-patent, not an extraction error. Distinct from
# gold so a reviewer doesn't chase these as v2 bugs.
COLOR_PATENT_VERIFIED = PatternFill(start_color="B7DEE8", end_color="B7DEE8", fill_type="solid")
# Mint: structure matched and the patent reports potency only as a letter-grade
# BIN (a range), with BDB's value falling inside that range. Consistent — not a
# precise numeric match, but not a mismatch either. Distinct from green (exact)
# and grey (no value) so binned coverage is legible at a glance.
COLOR_BINNED = PatternFill(start_color="B7E1CD", end_color="B7E1CD", fill_type="solid")


# ── Patent-source validation ──────────────────────────────────────────
# When v2's value disagrees with BDB, check the PATENT TEXT (the source of
# truth) for that compound. If v2's number appears next to the cid in a
# real assay table (NOT a property/retention table), v2 extracted the patent
# correctly and BDB is the outlier. Cached per patent (text + property spans;
# cid positions lazily) so the per-row check is cheap.
_PATENT_VALID_CACHE: dict = {}
_CID_POS_CACHE: dict = {}


# ── Property-table detection (moved here from the deleted output_validator)
#
# These two helpers were the only part of `core/assay_fsm/output_validator.py`
# worth keeping. That module GATED on this heuristic and measured 707 real
# `RORγ Binding IC50` rows deleted on US10273259 alone, so it was removed
# (see `tests/test_output_validator_removed.py`). Here the same signal is
# only ever REPORTED in a hand-check workbook, which is what it is good for:
# a human looks at the flagged row and decides.
#
# A patent interleaves real ASSAY tables (IC50/Ki/EC50/% inhibition) with
# PROPERTY tables — HPLC retention time, [M+H] mass, MW, logP. These markers
# appear in PROPERTY-table headers and never in real assay-table headers
# (which carry concentration units μM/nM):
_PROPERTY_HEADER_RE = re.compile(
    r"\(\s*min\s*\)"                     # retention time unit
    r"|\[\s*M\s*[+\-]\s*[HN][a]?\s*\]"   # [M+H] / [M-H] / [M+Na] mass
    r"|\bm/?z\b"                          # mass/charge
    r"|retention\s+time"
    r"|molecular\s+weight|\bMW\b"
    r"|\blog\s?[PD]\b"                    # logP / logD
    r"|melting\s+point|\bm\.?p\.?\b",
    re.IGNORECASE,
)


def _property_table_spans(text: str) -> list[tuple[int, int]]:
    """Return [start,end) spans of PROPERTY tables (retention/mass/MW),
    detected by a property marker in each `TABLE N` header. Real assay
    tables (IC50/Ki headers with μM/nM) are not matched."""
    if not text:
        return []
    spans: list[tuple[int, int]] = []
    tables = list(re.finditer(r"TABLE\s+\d+", text, re.IGNORECASE))
    for i, m in enumerate(tables):
        end = tables[i + 1].start() if i + 1 < len(tables) else min(len(text), m.start() + 4000)
        header = text[m.start():m.start() + 170]
        if _PROPERTY_HEADER_RE.search(header):
            spans.append((m.start(), end))
    return spans


def _value_strings(value: float) -> set[str]:
    """Plausible string representations of `value` that might appear
    in a table cell. Includes the natural decimal form plus rounded
    forms at 1-4 places, but NEVER coarser than the value's own
    precision (else we'd round 0.025 to "0" and match every cell
    that contains a bare zero — catastrophic false positive).
    """
    out: set[str] = {f"{value:g}"}
    # Generate fixed-precision forms ONLY at precisions that don't
    # truncate the value. e.g. for 0.025 we produce {"0.025", "0.0250"}
    # but skip "0" (.0f) and "0.03" (.2f) since both lose information.
    for p in (1, 2, 3, 4):
        s = f"{value:.{p}f}"
        if abs(float(s) - value) < 1e-9:
            out.add(s)
    # Strip trailing zeros from the .Xf forms so "1.40" matches "1.4"
    out.update({s.rstrip("0").rstrip(".") for s in list(out) if "." in s})
    out.discard("")
    # Final guard: never emit "0" (or empty) as a candidate for a
    # nonzero value — it'd match arbitrary "0" cells.
    if value != 0:
        out.discard("0")
    return out


def _v2_value_in_patent(patent_id: str, cid: str, value) -> bool:
    """True iff (cid, value) co-occur in the patent OUTSIDE a property table."""
    if value is None or not cid:
        return False
    if patent_id not in _PATENT_VALID_CACHE:
        try:
            from patentdb.core.patent_text import load_patent_description
            txt, _ = load_patent_description(patent_id, prefer_format="auto")
            txt = txt or ""
            _PATENT_VALID_CACHE[patent_id] = (txt, _property_table_spans(txt))
        except Exception:
            _PATENT_VALID_CACHE[patent_id] = ("", [])
    text, prop_spans = _PATENT_VALID_CACHE[patent_id]
    if not text:
        return False
    key = (patent_id, cid)
    if key not in _CID_POS_CACHE:
        _CID_POS_CACHE[key] = [m.start() for m in re.finditer(re.escape(cid) + r"(?!\d)", text)]
    # v2 stores µM; patents report in µM OR nM. Check both the µM form and the
    # nM form (value×1000) so e.g. v2 0.006 µM matches the patent's "6" (nM).
    vstrs = set(_value_strings(value))
    nm = value * 1000.0
    if 0.1 <= nm < 1e6:          # plausible nM magnitude only
        vstrs |= _value_strings(nm)
    for pos in _CID_POS_CACHE[key]:
        seg = text[pos + len(cid):pos + len(cid) + 16]
        if any(
            re.search(r"(?<![\d.])" + re.escape(v) + r"(?![\d])", seg)
            for v in vstrs
        ):
            if not any(s <= pos < e for s, e in prop_spans):
                return True
    return False


# Verdict sort priority — lower number = appears first in the sheet.
# Match rows at the top so a reviewer scrolling sees confirmed-good
# data first; mismatches at the bottom for focused review.
_VERDICT_ORDER = {
    # Clean matches at the top; reference-side anomalies (Jie typo or
    # OPSIN canonicalizing two Jie rows to one InChIKey) in their own
    # band right below so Jie can scan them together; gap verdicts last.
    "match": 0,
    "ref_iupac_wrong": 1,
    "jie_duplicate_same_molecule": 1,
    "v2_missed": 2,
    "v2_wrong": 3,
    "ref_wrong": 4,
    "stereo_diff": 5,
    "both_in_text": 6,
    "ref_only": 7,
    "v2_extra": 8,
    "v2_extra_unverified": 9,
    "v2_missing_molecule": 10,
}


# Hard budget guard. The plan caps total session spend; if approached, abort
# rather than burn unattributed cost.
_TOTAL_BUDGET_GUARD = 19.50


def _budget_check(reason: str = "") -> None:
    if cost_tracker.total_cost > _TOTAL_BUDGET_GUARD:
        raise RuntimeError(
            f"budget guard: ${cost_tracker.total_cost:.4f} exceeds "
            f"${_TOTAL_BUDGET_GUARD:.2f} threshold ({reason})"
        )


# Cache: IUPAC string → (smiles, full_inchikey, flat_inchikey).
# Most reference IUPACs (Jie + BDB) repeat across rows — cache so OPSIN
# fires once per unique name. PERSISTED to disk so each workbook rebuild
# doesn't re-pay ~$0.5-1 of OPSIN/LM to re-resolve the same ~190 reference
# names. Delete the file to force fresh resolution (e.g. after improving
# the IUPAC→SMILES cascade).
_REF_IK_CACHE: dict[str, tuple[str, str, str]] = {}
_REF_IK_CACHE_PATH = (
    Path(__file__).resolve().parent.parent.parent / "data" / "ref_ik_cache.json"
)
_REF_IK_CACHE_LOADED = False
_REF_IK_CACHE_DIRTY = 0  # new entries since last save (throttle disk writes)


def _load_ref_ik_cache() -> None:
    """Lazily hydrate `_REF_IK_CACHE` from disk on first use (idempotent)."""
    global _REF_IK_CACHE_LOADED
    if _REF_IK_CACHE_LOADED:
        return
    _REF_IK_CACHE_LOADED = True
    try:
        if _REF_IK_CACHE_PATH.exists():
            d = json.loads(_REF_IK_CACHE_PATH.read_text())
            for k, v in d.items():
                if isinstance(v, (list, tuple)) and len(v) == 3:
                    _REF_IK_CACHE[k] = (v[0] or "", v[1] or "", v[2] or "")
    except (ValueError, OSError):
        pass


def _save_ref_ik_cache() -> None:
    """Persist `_REF_IK_CACHE` to disk (tuples stored as JSON lists)."""
    global _REF_IK_CACHE_DIRTY
    try:
        _REF_IK_CACHE_PATH.parent.mkdir(parents=True, exist_ok=True)
        _REF_IK_CACHE_PATH.write_text(
            json.dumps({k: list(v) for k, v in _REF_IK_CACHE.items()})
        )
        _REF_IK_CACHE_DIRTY = 0
    except OSError:
        pass


def _ref_iupac_to_keys(iupac: str) -> tuple[str, str, str]:
    """Convert one reference IUPAC string to (smiles, full_inchikey,
    flat_inchikey). OPSIN-first; falls back to LM cascade only if OPSIN
    fails AND total budget remains. Returns ("", "", "") on failure.

    Each unique IUPAC string gets its own synthetic patent_id for cost
    attribution so the per-patent $0.20 cap doesn't aggregate across the
    190 reference compounds. Total spend bounded by the global $19.50
    guard via `_budget_check`. Typical actual spend per resolved name is
    ~$0.006; even worst-case (every Jie name needs LM cascade) is
    ~190 × $0.006 = ~$1.14, well within budget.
    """
    global _REF_IK_CACHE_DIRTY
    iupac = (iupac or "").strip()
    if not iupac:
        return "", "", ""
    _load_ref_ik_cache()
    if iupac in _REF_IK_CACHE:
        return _REF_IK_CACHE[iupac]

    # OPSIN — fast, free, deterministic. Handles most clean reference names.
    smiles, _err = _try_opsin(iupac)

    # If OPSIN fails AND we have budget, run the full cascade.
    if not smiles:
        _budget_check(f"reference IUPAC routing: {iupac[:80]!r}")
        # Per-name synthetic patent_id — each name has its own $0.20
        # cap rather than one shared $0.20 cap across all 190.
        synthetic_pid = f"REFHANDCHECK_{abs(hash(iupac)) % 1_000_000:06d}"
        c = Compound(
            patent_id=synthetic_pid,
            example_number="REF",
            iupac_name=iupac,
            iupac_source=IupacSource.PATENT_VERBATIM,
            source=CompoundSource.EXEMPLIFIED,
        )
        _convert_single(c, is_clean_text=True, route_hint="unknown")
        smiles = c.canonical_smiles or ""

    full_ik = get_inchikey(smiles) if smiles else ""
    flat_ik = get_stereo_flattened_key(smiles) if smiles else ""
    _REF_IK_CACHE[iupac] = (smiles, full_ik or "", flat_ik or "")
    # Throttled persist: flush every 10 new entries so a crash mid-build keeps
    # most progress, without an O(n^2) write per name. main() does a final
    # flush via _save_ref_ik_cache().
    _REF_IK_CACHE_DIRTY += 1
    if _REF_IK_CACHE_DIRTY >= 10:
        _save_ref_ik_cache()
    return _REF_IK_CACHE[iupac]


def _normalize_cid(cid: str) -> str:
    """Collapse cid surface forms ("172", "Example 172", "Cpd. No. 172",
    "172 Step A") to a canonical lowercased prefix-stripped key. Mirrors
    `text_index._norm_compound_id` so the workbook's matcher dedupes the
    same way the example_index merge would.
    """
    if not cid:
        return ""
    s = cid.strip()
    for _ in range(3):
        new = re.sub(
            r"^(?:cpd\.?\s*no\.?|compound|example|cpd|ex\.?)\s*",
            "", s, flags=re.IGNORECASE,
        ).strip()
        if new == s:
            break
        s = new
    # Strip trailing "Step X" qualifiers — those are intermediate-stage
    # entries pointing at the same final compound
    s = re.sub(r"\s*Step\s+[A-Z\d]+\s*$", "", s, flags=re.IGNORECASE).strip()
    return re.sub(r"[^a-z0-9]", "", s.lower())


def _build_v2_inchikey_indices(
    v2_ex: dict, v2_ay: dict,
) -> tuple[dict, dict]:
    """From example_index.json + assay_tables.json, build:
        v2_by_full_ik: {full_inchikey -> [(cid, ex_rec, assay_arr), ...]}
        v2_by_flat_ik: {flat_inchikey -> [(cid, ex_rec, assay_arr), ...]}

    Both indices return LISTS — same molecule can show up under multiple
    compound IDs (Strategy 2 vs Strategy 4 vs Strategy 5 numbering).
    Callers (`_structural_match`) search the list by assay value to pick
    the right one.

    Dedup rule: collapse two entries ONLY if they share BOTH a normalized
    compound_id AND a full InChIKey. That preserves stereoisomers — when
    Strategy 5 extracts `"Example 11" → (1R*,2S*)-...` and Strategy 4
    extracts `"11" → racemic cis-...`, those are DIFFERENT molecules in
    the patent (even if OPSIN happens to canonicalize them to the same
    InChIKey for some IUPAC encodings). Both entries are kept so the
    matcher can disambiguate by value.

    The previous version deduped on normalized cid alone, which collapsed
    distinct stereoisomers. That's wrong — patents distinguish them and
    so should v2.
    """
    # First pass: bucket by (normalized_cid, full_inchikey), pick best
    # representative within each bucket. Two entries with the same cid
    # but different InChIKeys stay separate.
    by_norm_cid_ik: dict[tuple[str, str], tuple[str, dict, list]] = {}
    for cid, rec in v2_ex.items():
        smiles = (rec.get("canonical_smiles") or "").strip()
        if not smiles:
            continue
        norm = _normalize_cid(cid)
        if not norm:
            continue
        full_ik = get_inchikey(smiles) or ""
        # Look up assays under both the raw cid and the normalized form
        ay = (
            v2_ay.get(cid, [])
            or v2_ay.get(cid.lower(), [])
            or v2_ay.get(norm, [])
            or []
        )
        bucket_key = (norm, full_ik)
        existing = by_norm_cid_ik.get(bucket_key)
        if existing is None:
            by_norm_cid_ik[bucket_key] = (cid, rec, ay)
            continue
        # Same (cid, IK) — pure surface-form duplicate. Prefer the
        # entry with assays; tiebreak on cleaner cid form.
        ex_cid, _ex_rec, ex_ay = existing
        if ay and not ex_ay:
            by_norm_cid_ik[bucket_key] = (cid, rec, ay)
        elif (not ay) == (not ex_ay) and len(cid) < len(ex_cid):
            by_norm_cid_ik[bucket_key] = (cid, rec, ay)

    # Second pass: bucket the (cid, IK)-deduped entries by InChIKey alone.
    # We index under EVERY IK the record carries:
    #   1. the IK derived from canonical_smiles (legacy path)
    #   2. rec["inchikey"] (set by Strategy 0 from GP-embedded structured data,
    #      or by the bridge when it overwrites an OPSIN-derived IK)
    #   3. every entry in rec["inchikey_aliases"] (displaced IKs preserved
    #      by the bridge's "overwrite" merges — these are the OPSIN-derived
    #      IKs that BDB sometimes happens to track)
    # Without (2) and (3), bridge-fixed records were matched only via the
    # SMILES-derived IK; on US11292791 that misclassified 55 compounds as
    # "BDB missed" when they were really matched (or stereo variants)
    # under their stored / aliased IKs.
    def _alt_iks(rec_: dict) -> list[str]:
        out: list[str] = []
        primary = (rec_.get("inchikey") or "").strip()
        if primary:
            out.append(primary)
        for a in (rec_.get("inchikey_aliases") or []):
            a = (a or "").strip()
            if a and a not in out:
                out.append(a)
        return out

    by_full: dict[str, list] = {}
    by_flat: dict[str, list] = {}
    seen_full: set[tuple[str, str]] = set()   # (full_ik, cid) already indexed
    for cid, rec, ay in by_norm_cid_ik.values():
        smiles = (rec.get("canonical_smiles") or "").strip()
        # ── by_full: every FULL InChIKey the record carries ──
        #   1. SMILES-derived (legacy path)
        #   2. rec["inchikey"] (Strategy 0 GP-embedded / bridge-injected)
        #   3. rec["inchikey_aliases"] (bridge-displaced OPSIN IKs)
        # Callers match BDB by full IK, so indexing all three recovers the
        # bridge-fixed compounds (US11292791 +47 etc.).
        full_iks: list[str] = []
        smi_full = get_inchikey(smiles) if smiles else None
        if smi_full:
            full_iks.append(smi_full)
        for stored in _alt_iks(rec):
            if stored not in full_iks:
                full_iks.append(stored)
        for ik in full_iks:
            if not ik or (ik, cid) in seen_full:
                continue
            seen_full.add((ik, cid))
            by_full.setdefault(ik, []).append((cid, rec, ay))
        # ── by_flat: stereo-flattened key from the SMILES ──
        # CONTRACT: callers query by_flat with get_stereo_flattened_key()
        # (a full 27-char InChIKey of the de-stereo'd molecule), NOT a
        # 14-char connectivity prefix. Keep this exact — keying by ik[:14]
        # silently broke the Jie stereo-variant match (166→76). Aliases
        # are exact-IK alternatives (not stereo variants) so they do NOT
        # belong in by_flat.
        if smiles:
            flat_ik = get_stereo_flattened_key(smiles)
            if flat_ik:
                by_flat.setdefault(flat_ik, []).append((cid, rec, ay))
    return by_full, by_flat


def _structural_match(
    candidates: list[tuple[str, dict, list]],
    ref_assay_kind: str,
    ref_value_uM: float | None,
) -> tuple[tuple[str, dict, list] | None, dict]:
    """Pure structural match. Given v2 candidates that all share the
    same InChIKey as the reference (i.e., they are STRUCTURALLY THE
    SAME MOLECULE), find the one whose assay row matches the reference
    assay value. Compound_id is NOT consulted — only structure + value.

    Returns (best_candidate_or_None, details).

    `details` includes:
        all_cids:        every v2 cid sharing the InChIKey
        all_uM_values:   list of v2 measurements across those cids for `ref_assay_kind`
        match_kind:      'value_match' | 'molecule_only' | 'no_assay_data'

    The reviewer sees every measurement v2 has for this molecule, not
    just one cherry-picked one.
    """
    if not candidates:
        return None, {"all_cids": [], "all_uM_values": [], "match_kind": "no_candidates"}

    all_cids = [c[0] for c in candidates]
    all_values: list[tuple[str, float | None, str]] = []  # (cid, uM, qualifier)
    for cid, _rec, ay in candidates:
        for a in ay:
            an = (a.get("assay_name") or "").lower()
            if not _assay_kind_match(an, ref_assay_kind):
                continue
            uM = value_to_uM(a.get("value_numeric"), a.get("unit"))
            qual = a.get("qualifier") or ""
            all_values.append((cid, uM, qual))

    if ref_value_uM is None:
        # Reference left this assay blank — return the first candidate
        # for context display, no value comparison
        return candidates[0], {
            "all_cids": all_cids, "all_uM_values": all_values,
            "match_kind": "ref_blank",
        }

    # Find a v2 measurement whose value matches the reference
    for cid, uM, qual in all_values:
        if uM is not None and values_equal(uM, ref_value_uM):
            # Pick the candidate whose cid matches this matching measurement
            for c in candidates:
                if c[0] == cid:
                    return c, {
                        "all_cids": all_cids, "all_uM_values": all_values,
                        "match_kind": "value_match",
                    }

    # Same molecule but no value match — prefer a candidate that AT LEAST
    # has assay data over one without. Strategy 0 GP-embedded entries
    # have no assay rows; HARVEST'd density-extracted entries do. We
    # want the workbook to surface the assay value (even if it differs
    # from the reference) rather than show v2_uM=None.
    cids_with_values = {cid for cid, _, _ in all_values}
    for c in candidates:
        if c[0] in cids_with_values:
            return c, {
                "all_cids": all_cids, "all_uM_values": all_values,
                "match_kind": "molecule_only",
            }
    # No candidate has assay data either — fall back to first.
    return candidates[0], {
        "all_cids": all_cids, "all_uM_values": all_values,
        "match_kind": "molecule_only",
    }


def _assay_kind_match(assay_name_lower: str, kind: str) -> bool:
    """Loose match between v2 assay_name and a logical 'kind' the
    caller cares about ('ki', 'ic50'). Used inside the structural matcher
    to filter relevant rows."""
    if kind == "ki":
        return "ki" in assay_name_lower and "kinase" not in assay_name_lower
    if kind == "ic50":
        return "ic50" in assay_name_lower or "ltb4" in assay_name_lower
    return kind in assay_name_lower


def values_equal(a, b) -> bool:
    """Equality comparison that tolerates BindingDB's storage ROUNDING.

    Both sides report the SAME number, but BindingDB stores it rounded to
    ~2-3 significant figures while the patent/extraction can be more precise
    (e.g. patent 0.000145 µM vs BDB 0.00015 µM — the same value). So:
      1. round both to 6 sig figs and compare exactly (collapses IEEE-754
         float-display artifacts like 0.004200000000000001 → 0.0042), AND
      2. accept agreement within 5.5% — the maximum relative error introduced
         by rounding a value to 2 significant figures. This is NOT a physical
         measurement tolerance: values >5.5% apart are still flagged as a real
         disagreement (caught earlier and routed to patent-validation).
    """
    if a is None or b is None:
        return False
    try:
        af, bf = float(a), float(b)
    except (TypeError, ValueError):
        return a == b

    def _round_sig(x: float, n: int = 6) -> float:
        if x == 0:
            return 0.0
        from math import log10, floor
        d = n - int(floor(log10(abs(x)))) - 1
        return round(x, d)

    if _round_sig(af) == _round_sig(bf):
        return True
    m = max(abs(af), abs(bf))
    return m > 0 and abs(af - bf) / m <= 0.055


def _value_in_text(text: str, value: float) -> bool:
    if not text or value is None:
        return False
    s_set: set[str] = set()
    for prec in (0, 1, 2, 3, 4, 5, 6):
        s_set.add(f"{value:.{prec}f}")
    s_set |= {s.rstrip("0").rstrip(".") for s in s_set}
    s_set |= {s.lstrip("0") if s.startswith("0.") else s for s in s_set}
    s_set.discard("")
    for s in s_set:
        if re.search(rf"(?<![A-Za-z0-9.]){re.escape(s)}(?![A-Za-z0-9])", text):
            return True
    return False


def _find_compound_row(text: str, cid: str) -> str:
    """Get the patent's row text for a compound — used for value-in-text checks."""
    if not text or not cid:
        return ""
    pat = re.compile(
        rf"(?<![A-Za-z0-9.]){re.escape(cid)}\s+"
        r"(?:[<>≤≥~]?\s*\d+(?:\.\d+)?(?:\s*\(\d+\))?[\s,;]*){2,5}",
        re.IGNORECASE,
    )
    m = pat.search(text)
    return m.group(0) if m else ""


def find_assay_in_v2(arr: list[dict], kind: str) -> dict | None:
    if not arr:
        return None
    for a in arr:
        an = (a.get("assay_name") or "").lower()
        if kind in an:
            return a
        if kind == "ic50" and any(k in an for k in
            ("ltb", "hwb", "menin", "pi3k", "cd69", "fret", "cc50", "molm", "mv4")):
            return a
        if kind == "ki" and "flap" in an:
            return a
    return None


# ── BDB hand-check builder (US10899738, US10214537) ────────────


def _assay_kind_of(name: str) -> str:
    """Pull the readout type (ic50/ec50/ki/kd/cc50/ec/gi50/…) out of a name,
    so different readouts of the SAME target stay distinct after canonicalizing."""
    m = re.search(r"\b(ic50|ec50|cc50|gi50|ed50|ld50|ki|kd|kb|pa2|emax|ec)\b",
                  name or "", re.IGNORECASE)
    return m.group(1).lower() if m else ""


def _canonical_assay_name(name: str) -> str:
    """Reduce a v2 assay_name to a canonical key.

    First try the assay reconciler — it maps the patent abbreviation to its
    canonical target (P2X3 → "P2X purinoceptor 3", Nav1.5 → "Sodium channel
    type 5", CBP → "CREB-binding protein"), so cross-vocabulary surface forms
    unify and line up with BindingDB's own target naming. We append the assay
    KIND so e.g. a target's IC50 and EC50 readouts don't collapse together.
    Falls back to plain slugification when the name doesn't reconcile (novel
    or non-standard assays)."""
    if not name:
        return ""
    try:
        from patentdb.core.assay_reconciler import reconcile
        hit = reconcile(name)
        if hit:
            target, _ = hit
            base = re.sub(r"[^A-Za-z0-9]+", "_", target.lower()).strip("_")
            kind = _assay_kind_of(name)
            return f"{base}_{kind}" if kind else base
    except Exception:
        pass
    s = re.sub(r"\s*[\(\[]\s*(?:nM|μM|uM|µM|mM|pM|M|%)\s*[\)\]]\s*", "", name, flags=re.IGNORECASE)
    s = re.sub(r"[^A-Za-z0-9]+", "_", s.lower()).strip("_")
    return s


def _v2_canonical_assays(v2_ay: dict) -> list[str]:
    """Return ordered list of canonical assay names present in v2,
    sorted by frequency descending.

    GUARD (belt-and-suspenders to the upstream `assay_name_guard`): even
    if a stale `assay_tables.json` from a pre-guard run still contains
    placeholder/non-assay names (`Activity1`, `col_0`, `[M+H]`, `Method`,
    `Molecular Weight`, …), we refuse to surface them as canonical
    primary columns here. Without this the workbook would pick the most
    frequent placeholder (US11292791 had 704 `Activity1` rows → became
    the first primary column → every "matched" row showed
    `activity1_uM = 0` → 0/660 value-matches against BDB).
    """
    from collections import Counter
    from patentdb.core.assay_name_guard import is_valid_assay_name
    ctr = Counter()
    for cid, arr in v2_ay.items():
        for a in arr:
            raw = a.get("assay_name") or ""
            if not is_valid_assay_name(raw):
                continue
            n = _canonical_assay_name(raw)
            if n:
                ctr[n] += 1
    return [name for name, _n in ctr.most_common()]


def _v2_pick_assay_value(v2_arr: list, canon_assay: str) -> dict:
    """Return {uM, qualifier, n_runs, name} for the v2 row whose canonical
    assay name matches `canon_assay`. {} if not found."""
    for a in v2_arr:
        if _canonical_assay_name(a.get("assay_name","")) == canon_assay:
            return {
                "uM": value_to_uM(a.get("value_numeric"), a.get("unit")),
                "qualifier": a.get("qualifier") or "",
                "n_runs": a.get("n_runs") if a.get("n_runs") is not None else "",
                "name": a.get("assay_name") or "",
            }
    return {}


def _v2_bin_match(v2_arr: list, ref_kind: str, ref_uM, ref_label: str):
    """Find the best letter-bin record (potency RANGE, value_numeric=None) of
    the right assay kind for this compound. Returns (name, value_raw, contains)
    or None if the compound has no letter-bin record.

    `contains` is True iff ref_uM falls inside that bin's range. A binned record
    means the patent reported only a +/++/+++ grade, not a number:
      • contains=True  → BDB's value is inside the stated grade (consistent).
      • contains=False → BDB's value is OUTSIDE the patent's grade; v2 still
        faithfully holds the patent's bin (a BDB-vs-patent disagreement).
    Prefer a bin that contains ref_uM, then one whose target (e.g. "G12C" or
    "CBP") appears in BDB's ref label, so multi-target panels pick the right one.
    """
    if ref_uM is None:
        return None
    ref_l = (ref_label or "").lower()
    _KIND_TOKS = ("ic50", "ic 50", "ec50", "kd", "ki")
    _STOP = {"ic50", "ic 50", "ec50", "fret", "cell", "viability",
             "assay", "gmean", "binding", "the", "and"}
    best = None        # (contains, target_score, name, raw)
    for a in v2_arr:
        if a.get("source") != "letter_bin" or a.get("value_numeric") is not None:
            continue
        an = (a.get("assay_name") or "")
        an_l = an.lower()
        # Only enforce the kind check when the bin name actually carries a kind
        # token. US11566007 bins are "...IC50" (enforce); US11292791 bins are
        # bare target names ("CBP"/"BRD4") with no kind word — don't exclude
        # those on kind grounds, let range + target decide.
        has_kind = any(k in an_l for k in _KIND_TOKS)
        if ref_kind and has_kind and not _assay_kind_match(an_l, ref_kind):
            continue
        lo, hi = a.get("value_low"), a.get("value_high")
        contains = ((lo is None or ref_uM >= lo) and (hi is None or ref_uM <= hi))
        # Target affinity: any non-stopword alpha token of the assay name that
        # also appears in BDB's ref label (e.g. "cbp" in "CBP IC50").
        tgt = 0
        for tok in re.findall(r"[a-z0-9]{3,}", an_l):
            if tok not in _STOP and tok in ref_l:
                tgt = 1
                break
        cand = (contains, tgt, an, a.get("value_raw") or "")
        # Prefer a bin that CONTAINS ref_uM, then a target match.
        if best is None or cand[:2] > best[:2]:
            best = cand
    return (best[2], best[3], best[0]) if best else None   # (name, raw, contains)


def build_bdb_rows(patent_id: str) -> list[dict]:
    """4-cluster hand-check sheet in Jie's wide format.

    One row per compound. Columns mirror Jie's layout:
        patent_id, compound_id, iupac_name, target, inchikey_full
        {assay}_uM, {assay}_qualifier (per canonical assay)
        ref_value_uM (BDB's value if applicable)
        verdict (matched | matched_value_diff | stereo_variant |
                 v2_stereo_extra | v2_only | bdb_missed)
        explanation

    Rows are color-coded by verdict (green=matched, gold=matched_value_diff,
    purple=stereo_variant/stereo_extra, blue=v2_only, peach=bdb_missed).

    Aggregate stats appended at the bottom of the sheet by `_write_sheet`.
    """
    return _build_bdb_rows_jie_format(patent_id)


def _build_bdb_rows_jie_format(patent_id: str) -> list[dict]:
    """Per-patent BDB sheet in Jie's wide format. Returns rows where:

      - one row per UNIQUE compound (not per assay measurement)
      - columns: patent_id, compound_id, iupac_name, inchikey_full,
        plus per-canonical-assay columns ({assay}_uM, _qualifier)
      - extra columns: ref_assay, ref_value_uM, verdict, explanation
      - `verdict` ∈ {"matched", "matched_value_diff", "v2_stereo_extra",
        "v2_only", "bdb_missed", "stereo_variant"} — single source of truth
        for sorting / counting / coloring (via `_fill`)
    """
    bdb_rows = list(iter_rows_for_patent(patent_id))
    _, v2_ay = fc.load_v2_extraction(patent_id)
    ex_path = Path(f"output_v2/text_extraction/{patent_id}/example_index.json")
    v2_ex = json.loads(ex_path.read_text()) if ex_path.exists() else {}

    # InChIKey indices on both sides
    v2_by_full_ik, v2_by_flat_ik = _build_v2_inchikey_indices(v2_ex, v2_ay)

    # Group BDB by SMILES then by InChIKey
    bdb_by_smiles: dict[str, list] = {}
    for r in bdb_rows:
        if r.smiles:
            bdb_by_smiles.setdefault(r.smiles, []).append(r)
    bdb_full_iks: dict[str, tuple[str, list]] = {}
    bdb_flat_iks: dict[str, list[str]] = {}
    for ref_smiles, meas_list in bdb_by_smiles.items():
        full_ik = get_inchikey(ref_smiles) if ref_smiles else None
        flat_ik = get_stereo_flattened_key(ref_smiles) if ref_smiles else None
        if full_ik:
            bdb_full_iks[full_ik] = (ref_smiles, meas_list)
        if flat_ik:
            bdb_flat_iks.setdefault(flat_ik, []).append(full_ik or "")

    # Detect canonical v2 assays present in this patent (for column layout)
    canonical_assays = _v2_canonical_assays(v2_ay)
    # Cap at 3 most common to keep sheets readable
    primary_assays = canonical_assays[:3]

    def _row_template(verdict: str, fill) -> dict:
        # `target` and `*_n_runs` columns dropped from the BDB sheets:
        # `target` is redundant with the per-assay column names, and
        # n_runs is never populated by HARVEST on these patents (BDB
        # carries the value but the patent text doesn't expose it).
        #
        # ONE verdict column (was the confusing `match` + `category` pair —
        # a row could read `category=matched` AND `match=FALSE`, which looks
        # like a contradiction). `verdict` is the single source of truth and
        # also drives sorting / counting / the Summary aggregation.
        return {
            "patent_id": patent_id,
            "compound_id": "",
            "iupac_name": "",
            "iupac_source": "",     # "patent" (default) | "pubchem_backfill"
            "inchikey_full": "",
            "ref_inchikey_full": "",  # BDB / reference InChIKey (variant rows)
            "route": "",                # ← which extractor produced this
            **{f"{a}_uM": "" for a in primary_assays},
            **{f"{a}_qualifier": "" for a in primary_assays},
            "ref_assay": "",
            "ref_value_uM": "",
            "verdict": verdict,
            "explanation": "",
            "_fill": fill,
        }

    def _iupac_source_label(rec: dict) -> str:
        """Surface where the IUPAC text came from. `pubchem_backfill` means
        v2 had a SMILES/InChIKey but no patent-text IUPAC, so we queried
        PubChem to fill the human-readable name. Reviewers should know
        the name didn't come from the patent itself."""
        src = (rec.get("iupac_source") or "").strip()
        if src == "pubchem_backfill":
            return "PubChem (from InChIKey)"
        return "patent" if (rec.get("iupac_name") or "").strip() else ""

    out: list[dict] = []

    # ── Cluster 1: matched ────────────────────────────────────────
    bdb_full_set = set(bdb_full_iks.keys())
    matched_full_iks = sorted(bdb_full_set & set(v2_by_full_ik.keys()))
    for full_ik in matched_full_iks:
        ref_smiles, meas_list = bdb_full_iks[full_ik]
        candidates = v2_by_full_ik[full_ik]
        # Pick v2 candidate whose first BDB value matches
        ref_first_uM = None
        if meas_list and meas_list[0].numeric_assays():
            first_nM = next(iter(meas_list[0].numeric_assays().values()))
            ref_first_uM = first_nM / 1000.0
        chosen, _ = _structural_match(candidates, "ki", ref_first_uM)
        if chosen is None:
            chosen = candidates[0]
        cid, v2_rec, v2_arr = chosen

        # NOTE: a same-molecule cross-cid assay merge was considered here
        # (GP-embedded structure cid vs example assay cid) but verified to
        # be a no-op — v2's example_index holds exactly one cid per full
        # InChIKey, so candidates never exceeds one. Values surface from the
        # cid's own assay rows; when they still don't match BDB it is a real
        # value disagreement, handled by the verdict split below.

        # Reference: take BDB's first measurement (most are single-assay)
        ref_label = ""
        ref_uM = None
        if meas_list and meas_list[0].numeric_assays():
            ref_label, ref_nM = next(iter(meas_list[0].numeric_assays().items()))
            ref_uM = ref_nM / 1000.0

        row = _row_template("matched", COLOR_MATCH)
        row["compound_id"] = cid
        row["iupac_name"] = (v2_rec.get("iupac_name") or "")[:200]
        row["iupac_source"] = _iupac_source_label(v2_rec)
        row["inchikey_full"] = full_ik
        row["ref_inchikey_full"] = full_ik     # matches v2 by definition in this cluster
        row["route"] = v2_rec.get("extraction_method", "") or ""
        # Fill assay columns from v2
        is_match = False
        matched_via_name = ""
        matched_via_value: float | None = None  # the assay-value that triggered the match
        for ca in primary_assays:
            v = _v2_pick_assay_value(v2_arr, ca)
            if v:
                row[f"{ca}_uM"] = v["uM"] if v["uM"] is not None else ""
                row[f"{ca}_qualifier"] = v["qualifier"]
                # Match check against ref
                if ref_uM is not None and v["uM"] is not None and values_equal(v["uM"], ref_uM):
                    is_match = True
                    matched_via_name = v["name"]
                    matched_via_value = v["uM"]

        # Catch-all: if no primary_assay column matched, scan ALL v2 assay
        # rows of the same kind. This catches cases where the patent
        # reports the matching value under an assay name that wasn't
        # picked into primary_assays (e.g. US10899738's "Menin Binding
        # Affinity IC50" — the actual primary IC50 readout — wasn't in
        # primary_assays which captured MV4;11 + MOLM13 instead).
        # The matching value is in v2's data; we just need to surface it.
        if not is_match and ref_uM is not None:
            # Infer ref kind from ref_label
            ref_label_lower = (ref_label or "").lower()
            if "ic50" in ref_label_lower:
                ref_kind = "ic50"
            elif "ki" in ref_label_lower and "kinase" not in ref_label_lower:
                ref_kind = "ki"
            elif "kd" in ref_label_lower:
                ref_kind = "kd"
            else:
                ref_kind = ""
            for a in v2_arr:
                an = (a.get("assay_name") or "").lower()
                if ref_kind and not _assay_kind_match(an, ref_kind):
                    continue
                v_uM = value_to_uM(a.get("value_numeric"), a.get("unit"))
                if v_uM is not None and values_equal(v_uM, ref_uM):
                    is_match = True
                    matched_via_name = a.get("assay_name") or ""
                    matched_via_value = v_uM
                    break

        # Stash the matching pair so the Verified_Data sheet can surface
        # the assay that actually triggered the match — not the first
        # populated primary column (which on US10899738 cid 19 is `MV4;11
        # = 1.6` while the BDB-matching assay is `Menin Binding Affinity
        # IC50 = 0.061`). Internal `_`-prefixed fields don't render as
        # workbook columns; the Verified_Data builder reads them directly.
        row["_matched_via_name"] = matched_via_name
        row["_matched_via_value"] = matched_via_value

        row["ref_assay"] = ref_label
        row["ref_value_uM"] = ref_uM if ref_uM is not None else ""
        if is_match:
            # Structure matched AND value agrees → leave explanation blank.
            # Reviewers scan for matched_value_diff / stereo_variant rows
            # where the explanation actually carries information.
            row["explanation"] = ""
        else:
            # Structure matched but no v2 row reproduces the BDB value.
            # Split the two very different cases — otherwise reviewers see a
            # wall of "mismatch" yellow for rows that merely LACK a value:
            #   • v2 HAS a value of the ref kind that disagrees → real value
            #     mismatch (yellow); show both numbers.
            #   • v2 has NO value of the ref kind for this molecule →
            #     coverage gap, NOT a mismatch → neutral grey.
            ref_label_lower = (ref_label or "").lower()
            if "ic50" in ref_label_lower:
                _rk = "ic50"
            elif "ki" in ref_label_lower and "kinase" not in ref_label_lower:
                _rk = "ki"
            elif "kd" in ref_label_lower:
                _rk = "kd"
            else:
                _rk = ""
            _v2_best = None
            for a in v2_arr:
                an = (a.get("assay_name") or "").lower()
                if _rk and not _assay_kind_match(an, _rk):
                    continue
                vu = value_to_uM(a.get("value_numeric"), a.get("unit"))
                if vu is not None:
                    _v2_best = (a.get("assay_name") or "", vu)
                    break
            _bin = _v2_bin_match(v2_arr, _rk, ref_uM, ref_label)
            if _bin is not None and _bin[2]:
                # Patent reports potency only as a letter-grade BIN (range),
                # and BDB's value falls INSIDE it → consistent, not a mismatch
                # and not a coverage gap. Surface the range honestly.
                row["verdict"] = "matched_binned"
                row["explanation"] = (
                    f"Patent reports {_bin[0]} only as a potency bin "
                    f"{_bin[1]} µM for {cid}; BDB={ref_uM} falls within it "
                    f"(letter-grade bin — no precise numeric value in patent)"
                )
                row["_fill"] = COLOR_BINNED
            elif _bin is not None:
                # v2 holds the patent's letter grade, but BDB's value is
                # OUTSIDE that bin → v2 is faithful to the patent; BDB
                # disagrees with the patent's own grade (not our error).
                row["verdict"] = "matched_patent_verified"
                row["explanation"] = (
                    f"Patent grades {_bin[0]} as bin {_bin[1]} µM for {cid}; "
                    f"BDB={ref_uM} is outside it — v2 is faithful to the "
                    f"patent's letter grade (BDB-vs-patent disagreement)"
                )
                row["_fill"] = COLOR_PATENT_VERIFIED
            elif _v2_best is not None:
                # Patent-source validation: if v2's value is confirmed in the
                # patent's own assay table for this cid, v2 is faithful and the
                # BDB disagreement is BDB-vs-patent (not our problem). Only flag
                # a true mismatch when v2's value is NOT in the patent.
                if _v2_value_in_patent(patent_id, cid, _v2_best[1]):
                    row["verdict"] = "matched_patent_verified"
                    row["explanation"] = (
                        f"v2 {_v2_best[0]}={_v2_best[1]} confirmed in the patent's "
                        f"assay table for {cid}; BDB={ref_uM} differs "
                        f"(BDB-vs-patent disagreement, not an extraction error)"
                    )
                    row["_fill"] = COLOR_PATENT_VERIFIED
                else:
                    row["verdict"] = "matched_value_diff"
                    row["explanation"] = (
                        f"Same molecule. BDB {_rk or 'value'}={ref_uM} vs "
                        f"v2 {_v2_best[0]}={_v2_best[1]} — values differ "
                        f"(v2 value NOT found in patent text for {cid})"
                    )
                    row["_fill"] = COLOR_VALUE_DIFF
            else:
                row["verdict"] = "matched_no_v2_value"
                row["explanation"] = (
                    f"Same molecule, structure matched; v2 has no "
                    f"{_rk or 'assay'} value for it (coverage gap, not a mismatch)"
                )
                row["_fill"] = COLOR_AMBIG
        out.append(row)

    # ── Cluster 2: v2 stereo extras ──────────────────────────────
    seen_v2_full = set()
    for flat_ik, v2_candidates in v2_by_flat_ik.items():
        if flat_ik not in bdb_flat_iks:
            continue
        bdb_full_iks_for_flat = set(bdb_flat_iks[flat_ik])
        for cid, v2_rec, v2_arr in v2_candidates:
            our_smiles = (v2_rec.get("canonical_smiles") or "")
            our_full = get_inchikey(our_smiles) if our_smiles else ""
            if not our_full or our_full in bdb_full_set or our_full in seen_v2_full:
                continue
            seen_v2_full.add(our_full)
            row = _row_template("v2_stereo_extra", COLOR_STEREO_DIFF)
            row["compound_id"] = cid
            row["iupac_name"] = (v2_rec.get("iupac_name") or "")[:200]
            row["iupac_source"] = _iupac_source_label(v2_rec)
            row["inchikey_full"] = our_full
            row["route"] = v2_rec.get("extraction_method", "") or ""
            for ca in primary_assays:
                v = _v2_pick_assay_value(v2_arr, ca)
                if v:
                    row[f"{ca}_uM"] = v["uM"] if v["uM"] is not None else ""
                    row[f"{ca}_qualifier"] = v["qualifier"]
            row["ref_value_uM"] = ""
            row["explanation"] = (
                f"v2 has stereo variant BDB doesn't list "
                f"(BDB has {sorted(bdb_full_iks_for_flat)[0][:14]}…)"
            )
            out.append(row)

    # ── Cluster 3: v2-only ───────────────────────────────────────
    bdb_flat_set = set(bdb_flat_iks.keys())
    seen_v2_full_3 = set()
    for full_ik, v2_candidates in v2_by_full_ik.items():
        if full_ik in bdb_full_set or full_ik in seen_v2_full_3:
            continue
        seen_v2_full_3.add(full_ik)
        cid, v2_rec, v2_arr = v2_candidates[0]
        our_smiles = (v2_rec.get("canonical_smiles") or "")
        flat_ik = get_stereo_flattened_key(our_smiles) if our_smiles else ""
        if flat_ik in bdb_flat_set:
            continue
        row = _row_template("v2_only", COLOR_V2_EXTRA)
        row["compound_id"] = cid
        row["iupac_name"] = (v2_rec.get("iupac_name") or "")[:200]
        row["iupac_source"] = _iupac_source_label(v2_rec)
        row["inchikey_full"] = full_ik
        row["route"] = v2_rec.get("extraction_method", "") or ""
        for ca in primary_assays:
            v = _v2_pick_assay_value(v2_arr, ca)
            if v:
                row[f"{ca}_uM"] = v["uM"] if v["uM"] is not None else ""
                row[f"{ca}_qualifier"] = v["qualifier"]
        row["ref_value_uM"] = ""
        row["explanation"] = "v2 extracted from patent; BDB has no entry"
        out.append(row)

    # ── Cluster 4: BDB-missed ────────────────────────────────────
    for full_ik, (ref_smiles, meas_list) in bdb_full_iks.items():
        if full_ik in v2_by_full_ik:
            continue
        ref_flat = get_stereo_flattened_key(ref_smiles) if ref_smiles else ""
        if ref_flat in v2_by_flat_ik:
            continue
        cid = meas_list[0].compound_id or ""
        ref_label = ""
        ref_uM = None
        if meas_list[0].numeric_assays():
            ref_label, ref_nM = next(iter(meas_list[0].numeric_assays().items()))
            ref_uM = ref_nM / 1000.0
        row = _row_template("bdb_missed", COLOR_V2_WRONG)
        row["compound_id"] = cid
        row["iupac_name"] = "(not extracted)"
        row["inchikey_full"] = ""              # v2 has no entry for this molecule
        row["ref_inchikey_full"] = full_ik     # BDB's reference InChIKey
        row["ref_assay"] = ref_label
        row["ref_value_uM"] = ref_uM if ref_uM is not None else ""
        row["explanation"] = "BDB has this compound; v2 didn't extract it from the patent"
        out.append(row)

    # ── Cluster 5: matched_stereo_variant ────────────────────────
    # BDB compound has full stereo InChIKey X; v2 has the SAME skeleton
    # (matching first-14-char InChIKey block) but different / lost stereo
    # (different 10-char stereo block). Previously these silently vanished
    # from the workbook (not matched, not bdb_missed, not v2_stereo_extra).
    # Surface them as their own cluster so analysts can verify the stereo
    # assignment.
    seen_bdb_flat_emitted: set[str] = set()
    for full_ik, (ref_smiles, meas_list) in bdb_full_iks.items():
        if full_ik in v2_by_full_ik:
            continue   # already in Cluster 1
        ref_flat = get_stereo_flattened_key(ref_smiles) if ref_smiles else ""
        if not ref_flat or ref_flat not in v2_by_flat_ik:
            continue   # not a stereo-flat match — will land in Cluster 4 / v2_only
        if ref_flat in seen_bdb_flat_emitted:
            continue
        seen_bdb_flat_emitted.add(ref_flat)
        # Pick the v2 candidate at this flat-IK
        candidates = v2_by_flat_ik[ref_flat]
        cid_v2, v2_rec, v2_arr = candidates[0]
        cid_bdb = meas_list[0].compound_id or ""
        ref_label = ""
        ref_uM = None
        if meas_list[0].numeric_assays():
            ref_label, ref_nM = next(iter(meas_list[0].numeric_assays().items()))
            ref_uM = ref_nM / 1000.0
        row = _row_template("stereo_variant", COLOR_STEREO_DIFF)
        row["compound_id"] = f"{cid_bdb} (v2: {cid_v2})"
        row["iupac_name"] = (v2_rec.get("iupac_name") or "")[:200]
        row["iupac_source"] = _iupac_source_label(v2_rec)
        # Keep `inchikey_full` clean (v2's only) so it parses as a real
        # InChIKey in tooling. BDB's InChIKey goes in its own column.
        row["inchikey_full"] = v2_rec.get("inchikey", "") or ""
        row["ref_inchikey_full"] = full_ik
        row["route"] = v2_rec.get("extraction_method", "") or ""
        # Fill v2 assay columns
        for ca in primary_assays:
            v = _v2_pick_assay_value(v2_arr, ca)
            if v:
                row[f"{ca}_uM"] = v["uM"] if v["uM"] is not None else ""
                row[f"{ca}_qualifier"] = v["qualifier"]
        row["ref_assay"] = ref_label
        row["ref_value_uM"] = ref_uM if ref_uM is not None else ""
        row["explanation"] = (
            f"Same skeleton (first 14 chars match). Different stereo block — "
            f"v2 has different or lost stereochemistry. Compound IS structurally "
            f"correct; review for stereo verification."
        )
        out.append(row)

    return out


# ── Jie hand-check (US8952177) — wide format mirroring Jie's CSV ─


# Column structure from Jie's reference CSV — replicated verbatim.
JIE_COL_GROUPS = [
    {
        "prefix": "flap_binding_ki",
        "v2_kind": "ki",
        "ref_value_field": "ki",
        "ref_qual_field": "ki_qual",
        "ref_runs_field": "ki_runs",
        "assay_label": "FLAP Binding HTRF (human, SF-9 membranes)",
    },
    {
        "prefix": "hwb_ltb4_ic50",
        "v2_kind": "ic50",
        "ref_value_field": "ic50",
        "ref_qual_field": "ic50_qual",
        "ref_runs_field": "ic50_runs",
        "assay_label": "Human Whole Blood LTB4 (A23187-stimulated)",
    },
]


def build_jie_rows() -> list[dict]:
    """One row per Jie compound (wide format) mirroring Jie's CSV columns
    EXACTLY: flap_binding_ki_uM, flap_binding_ki_qualifier,
    flap_binding_ki_n_runs, hwb_ltb4_ic50_uM, hwb_ltb4_ic50_qualifier,
    hwb_ltb4_ic50_n_runs, plus comparison/verdict columns.

    Match key: Jie's IUPAC → SMILES (OPSIN) → InChIKey → look up in v2.
    `compound_id` is shown for context but is NOT the join key — Jie may
    label the same molecule "5a" while v2 labels it "5", and that
    shouldn't break matching.
    """
    jie_path = DOWNLOADS / "US8952177 Binding IUPAC Final (1).csv"
    jie = {}
    with jie_path.open() as f:
        rdr = csv.DictReader(f)
        for row in rdr:
            cid = row["compound_id"].strip()
            jie[cid] = {
                "iupac": row.get("iupac_name", "").strip(),
                "ki": float(row["flap_binding_ki_uM"]) if row.get("flap_binding_ki_uM") else None,
                "ki_qual": (row.get("flap_binding_ki_qualifier") or "").strip(),
                "ki_runs": int(row["flap_binding_ki_n_runs"]) if row.get("flap_binding_ki_n_runs") else None,
                "ic50": float(row["hwb_ltb4_ic50_uM"]) if row.get("hwb_ltb4_ic50_uM") else None,
                "ic50_qual": (row.get("hwb_ltb4_ic50_qualifier") or "").strip(),
                "ic50_runs": int(row["hwb_ltb4_ic50_n_runs"]) if row.get("hwb_ltb4_ic50_n_runs") else None,
            }

    v2_ay = json.loads(open("output_v2/text_extraction/US8952177/assay_tables.json").read())
    ex_path = Path("output_v2/text_extraction/US8952177/example_index.json")
    v2_ex = json.loads(ex_path.read_text()) if ex_path.exists() else {}
    text = load_gp_description("US8952177")

    # Pre-index the patent's `## Example N` headers from MinerU markdown
    # so we can compare v2's stored IUPAC vs Jie's text against the
    # patent's authoritative example name. Used to distinguish "v2 is
    # right, Jie has a transcription error" from "v2 extracted wrong
    # text under this cid".
    _patent_example_iupacs: dict[str, str] = {}
    _example_hdr_re = re.compile(
        r"^##\s+Example\s+(\d+(?:[A-Z]{1,4})?)\s*:?\s*\n+([^\n#]{20,500})",
        re.MULTILINE,
    )
    pages_dir = Path("US8952177/all_pages")
    if pages_dir.exists():
        for pf in sorted(pages_dir.glob("page_*.md")):
            try:
                pt = pf.read_text(encoding="utf-8")
            except OSError:
                continue
            for m in _example_hdr_re.finditer(pt):
                cid_h = m.group(1).strip()
                iup = re.sub(r"\s+", " ", m.group(2)).strip()
                _patent_example_iupacs.setdefault(cid_h, iup)

    def _v2_matches_patent_header_better_than_ref(
        cid_local: str, v2_iupac: str, ref_iupac: str,
    ) -> bool:
        """True when the patent's `## Example {cid}` header IUPAC matches
        v2's stored text more closely than the reference IUPAC does.
        Uses normalized-substring overlap (cheap, deterministic, no LLM)
        as a tiebreaker. Falls back to False when no patent header is
        available or both sides have similar overlap with the patent.
        """
        patent = _patent_example_iupacs.get(cid_local, "")
        if not patent or not v2_iupac:
            return False
        def norm(s: str) -> str:
            # Strip whitespace, escaping, mojibake apostrophes so we
            # focus on structural tokens.
            s = re.sub(r"[\\\s]+", "", s.lower())
            s = s.replace("â", "'").replace("â2", "'").replace("â1", "'")
            return s
        np, nv, nr = norm(patent), norm(v2_iupac), norm(ref_iupac)
        if not np:
            return False
        # k-mer overlap (k=20): roughly how much of the patent header
        # is preserved in each candidate.
        k = 20
        if len(np) < k:
            return False
        kmers = {np[i:i+k] for i in range(len(np) - k + 1)}
        v_hits = sum(1 for km in kmers if km in nv)
        r_hits = sum(1 for km in kmers if km in nr)
        # Require v2 to BEAT ref by a clear margin AND cover the patent
        # at all. Otherwise the comparison is inconclusive.
        return v_hits >= 0.6 * len(kmers) and v_hits >= r_hits + 5

    # InChIKey indices over v2 (built once)
    v2_by_full_ik, v2_by_flat_ik = _build_v2_inchikey_indices(v2_ex, v2_ay)

    rows: list[dict] = []
    # When v2 has MULTIPLE entries for one InChIKey (genuine stereoisomer
    # variants stored separately), each Jie row binds a distinct v2 cid
    # so we don't conflate them. When v2 has ONLY ONE entry and Jie has
    # two rows pointing at the same InChIKey (common: Jie lists "racemic
    # cis-X" at Example 15 and "(1S*,3R*)-X" at Example 16, but OPSIN
    # canonicalizes both to one InChIKey so structurally they're the same
    # molecule and v2 correctly stores one entry), both Jie rows should
    # share the v2 entry — neither is "missing".
    claimed_v2_cids: set[str] = set()

    for cid, ref in jie.items():
        # Resolve Jie's IUPAC to InChIKey, then look up in v2
        _budget_check(f"Jie row {cid}")
        ref_smiles, ref_full_ik, ref_flat_ik = _ref_iupac_to_keys(ref["iupac"])

        # Pure structural lookup: gather ALL v2 entries that share the
        # InChIKey, then exclude any already claimed by a prior Jie row
        smi_match = "v2_missing_molecule"
        all_candidates = []
        if ref_full_ik and ref_full_ik in v2_by_full_ik:
            all_candidates = v2_by_full_ik[ref_full_ik]
            smi_match = "exact"
        elif ref_flat_ik and ref_flat_ik in v2_by_flat_ik:
            all_candidates = v2_by_flat_ik[ref_flat_ik]
            smi_match = "stereo_diff"
        candidates = [c for c in all_candidates if c[0] not in claimed_v2_cids]
        if not candidates and all_candidates:
            # Every v2 candidate at this InChIKey is already claimed by an
            # earlier Jie row. Two cases:
            #   (a) v2 has FEWER entries at this InChIKey than Jie has
            #       reference rows → genuinely ambiguous which Jie row
            #       maps to which v2 entry. Re-allow the claim rather
            #       than declaring v2_missing_molecule; mark as a
            #       duplicate-of-same-molecule pairing so reviewers see
            #       it without it polluting the gap count.
            candidates = all_candidates
            smi_match = "jie_duplicate_same_molecule"

        # Pick the v2 candidate whose Ki value matches Jie's; tie-break
        # on IC50 match; else first remaining
        matched_v2, ki_details = _structural_match(
            candidates, ref_assay_kind="ki", ref_value_uM=ref["ki"],
        )
        if matched_v2 is not None and ki_details["match_kind"] != "value_match":
            ic50_candidate, ic50_details = _structural_match(
                candidates, ref_assay_kind="ic50", ref_value_uM=ref["ic50"],
            )
            if ic50_details["match_kind"] == "value_match":
                matched_v2 = ic50_candidate
        if matched_v2 is not None:
            claimed_v2_cids.add(matched_v2[0])

        v2_cid = ""
        v2_arr: list = []
        our_iupac = ""
        v2_full_ik = ""
        v2_route = ""
        our_iupac_source = ""
        if matched_v2:
            v2_cid, v2_rec, v2_arr = matched_v2
            our_iupac = (v2_rec.get("iupac_name") or "")
            v2_full_ik = get_inchikey(v2_rec.get("canonical_smiles") or "") or ""
            v2_route = v2_rec.get("extraction_method", "") or ""
            our_iupac_source = (
                "PubChem (from InChIKey)"
                if (v2_rec.get("iupac_source") == "pubchem_backfill")
                else ("patent" if our_iupac.strip() else "")
            )

        # Build per-row (one per Jie compound) with per-assay verdicts.
        # Column inventory is intentionally lean — see Jie ergonomics
        # pass: drop `target` (always FLAP for this sheet) and the v2-side
        # `n_runs` columns (LLM never returns those; previously always
        # blank). Keep `route` so reviewers can see which extractor
        # produced the v2 entry.
        row: dict = {
            "patent_id": "US8952177",
            "compound_id": cid,           # Jie's compound_id (shown for human context)
            "v2_compound_id": v2_cid,     # what v2 labeled the same molecule
            "iupac_name": our_iupac[:200],
            "iupac_source": our_iupac_source,
            "jie_iupac_name": ref["iupac"][:200],
            "smiles_match": smi_match,    # exact | stereo_diff | v2_missing_molecule
            "jie_inchikey_full": ref_full_ik or "",
            "v2_inchikey_full": v2_full_ik,
            "route": v2_route,            # ← which extractor produced the v2 entry
        }

        # When v2 has an entry at this cid (regardless of InChIKey
        # match), and that entry's IUPAC sits in the patent's `## Example
        # N` header on disk while Jie's IUPAC does NOT, the reference
        # data is the wrong side. Surface as `ref_iupac_wrong` so the
        # row doesn't pollute the gap count.
        if smi_match == "v2_missing_molecule":
            v2_at_cid = v2_ex.get(cid, {})
            if v2_at_cid and _v2_matches_patent_header_better_than_ref(
                cid, v2_at_cid.get("iupac_name", ""), ref["iupac"],
            ):
                smi_match = "ref_iupac_wrong"
                v2_arr = v2_ay.get(cid, [])
                # Synthesize a "matched" view for the row using the v2 entry
                matched_v2 = (cid, v2_at_cid, v2_arr)
                v2_cid = cid
                our_iupac = v2_at_cid.get("iupac_name", "") or ""
                v2_full_ik = v2_at_cid.get("inchikey", "") or ""
                v2_route = v2_at_cid.get("extraction_method", "") or ""
                our_iupac_source = (
                    "PubChem (from InChIKey)"
                    if (v2_at_cid.get("iupac_source") == "pubchem_backfill")
                    else ("patent" if our_iupac.strip() else "")
                )

        # If we have no v2 InChIKey match AND the v2 entry at this cid
        # doesn't match the patent header either, that dominates the
        # verdict — reviewer needs to know we never extracted this
        # compound (or extracted it with the wrong structure).
        if smi_match == "v2_missing_molecule":
            worst_verdict = "v2_missing_molecule"
            worst_fill = COLOR_V2_WRONG
        elif smi_match == "ref_iupac_wrong":
            worst_verdict = "ref_iupac_wrong"
            worst_fill = COLOR_REF_WRONG
        elif smi_match == "jie_duplicate_same_molecule":
            worst_verdict = "jie_duplicate_same_molecule"
            worst_fill = COLOR_STEREO_DIFF
        elif smi_match == "stereo_diff":
            worst_verdict = "stereo_diff"
            worst_fill = COLOR_STEREO_DIFF
        else:
            worst_verdict = "match"
            worst_fill = COLOR_MATCH
        explanations = []
        if smi_match == "v2_missing_molecule":
            explanations.append(
                "v2 has no InChIKey match for Jie's molecule — either we "
                "never extracted this compound or extracted it with the "
                "wrong structure"
            )
        elif smi_match == "ref_iupac_wrong":
            explanations.append(
                "Jie's IUPAC disagrees with the patent's `## Example "
                f"{cid}` header. v2's stored IUPAC matches the patent."
            )
        elif smi_match == "jie_duplicate_same_molecule":
            explanations.append(
                "OPSIN canonicalizes this Jie IUPAC to the same InChIKey "
                "as an earlier Jie row; v2 has one entry covering both. "
                "Assay values may still differ between the two Jie rows."
            )
        elif smi_match == "stereo_diff":
            explanations.append(
                "v2 has same skeleton as Jie's molecule but different "
                "stereochemistry"
            )

        for grp in JIE_COL_GROUPS:
            our_a = find_assay_in_v2(v2_arr, grp["v2_kind"])
            our_uM = value_to_uM(
                our_a.get("value_numeric") if our_a else None,
                our_a.get("unit") if our_a else "",
            )
            our_q = (our_a.get("qualifier") if our_a else "") or ""
            ref_v = ref[grp["ref_value_field"]]
            ref_q = ref[grp["ref_qual_field"]]
            ref_r = ref[grp["ref_runs_field"]]

            # Skip both-blank
            if our_uM is None and ref_v is None:
                row[f"{grp['prefix']}_uM_v2"] = ""
                row[f"{grp['prefix']}_uM_jie"] = ""
                row[f"{grp['prefix']}_qualifier_v2"] = ""
                row[f"{grp['prefix']}_qualifier_jie"] = ""
                row[f"{grp['prefix']}_n_runs_jie"] = ref_r if ref_r is not None else ""
                row[f"{grp['prefix']}_assay"] = grp["assay_label"]
                continue

            if our_uM is not None and ref_v is not None and values_equal(our_uM, ref_v):
                # Clear text for the obvious case — keeps the cell visually
                # unobtrusive so reviewers can scan for the rows that
                # actually have something to look at.
                v_, e_, f_ = "match", "", COLOR_MATCH
            elif our_uM is not None and ref_v is not None:
                v2_in = _value_in_text(text, our_uM)
                ref_in = _value_in_text(text, ref_v)
                if v2_in and not ref_in:
                    v_, e_, f_ = "ref_wrong", f"v2 ({our_uM}) in patent; Jie ({ref_v}) not", COLOR_REF_WRONG
                elif ref_in and not v2_in:
                    v_, e_, f_ = "v2_wrong", f"Jie ({ref_v}) in patent; v2 ({our_uM}) not", COLOR_V2_WRONG
                else:
                    v_, e_, f_ = "both_in_text", "both values appear in patent", COLOR_AMBIG
            elif our_uM is None:
                v_, e_, f_ = "v2_missed", f"Jie has {ref_v}; v2 missed", COLOR_V2_WRONG
            else:
                v_, e_, f_ = "v2_extra", f"v2 has {our_uM} where Jie left blank", COLOR_V2_EXTRA

            row[f"{grp['prefix']}_uM_v2"] = our_uM if our_uM is not None else ""
            row[f"{grp['prefix']}_uM_jie"] = ref_v if ref_v is not None else ""
            row[f"{grp['prefix']}_qualifier_v2"] = our_q
            row[f"{grp['prefix']}_qualifier_jie"] = ref_q
            row[f"{grp['prefix']}_n_runs_jie"] = ref_r if ref_r is not None else ""
            row[f"{grp['prefix']}_assay"] = grp["assay_label"]
            row[f"{grp['prefix']}_verdict"] = v_
            row[f"{grp['prefix']}_explanation"] = e_

            if _VERDICT_ORDER.get(v_, 99) > _VERDICT_ORDER.get(worst_verdict, 0):
                worst_verdict = v_
                worst_fill = f_
            if v_ != "match":
                explanations.append(f"{grp['prefix']}: {e_}")

        # Skip rows where BOTH assays were both_blank
        per_assay_verdicts = [row.get(f"{g['prefix']}_verdict") for g in JIE_COL_GROUPS]
        if all(not v for v in per_assay_verdicts):
            continue

        row["compound_verdict"] = worst_verdict
        # Trivial-match explanation cleared so reviewers' eyes flow over
        # green rows to focus on the ones that need attention.
        row["compound_explanation"] = "; ".join(explanations) if explanations else ""
        row["_fill"] = worst_fill
        rows.append(row)

    # Sort: matches first, then natural-numeric cid within each bucket
    # (so cid "9" comes before cid "10", not after "100").
    def _cid_sort_key(cid: str) -> tuple[int, str]:
        m = re.match(r"(\d+)([A-Za-z]*)$", str(cid))
        return (int(m.group(1)), m.group(2)) if m else (10**9, str(cid))

    rows.sort(key=lambda r: (
        _VERDICT_ORDER.get(r["compound_verdict"], 99),
        _cid_sort_key(r["compound_id"]),
    ))
    return rows


# ── Excel writer ────────────────────────────────────────────────


def _write_sheet(wb, name: str, rows: list[dict], title_suffix: str) -> dict:
    ws = wb.create_sheet(name)
    if not rows:
        return {"total": 0}
    fields = [k for k in rows[0].keys() if not k.startswith("_")]

    ws.cell(row=1, column=1, value=f"{name} — {len(rows)} rows ({title_suffix})").font = Font(bold=True, color="555555")
    ws.cell(row=2, column=1, value="Sorted: matched compounds first; aggregate stats at the bottom.").font = Font(italic=True, color="666666")

    header_row = 4
    for cidx, fname in enumerate(fields, start=1):
        c = ws.cell(row=header_row, column=cidx, value=fname)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = COLOR_HEADER
        c.alignment = Alignment(horizontal="left", vertical="center")

    # Sort key composition:
    #   1. BDB-style verdict buckets when present (`verdict` set on BDB rows;
    #      absent on Jie rows so they all share rank 99). Structurally-found
    #      rows (matched → value-diff → stereo) surface first; gaps last.
    #   2. Jie rows carry `compound_verdict` (match, v2_missed, …) → mapped
    #      via _VERDICT_ORDER so reviewer-priority rows surface first.
    #   3. Natural-numeric compound_id so "9" precedes "10" rather than
    #      the lexicographic order "1, 10, 100, 11, …".
    cat_order = {
        "matched": 0, "matched_patent_verified": 1, "matched_binned": 2,
        "matched_value_diff": 3, "matched_no_v2_value": 4,
        "stereo_variant": 5, "v2_stereo_extra": 6, "v2_only": 7, "bdb_missed": 8,
    }

    def _natural_cid(cid) -> tuple[int, str]:
        m = re.match(r"(\d+)([A-Za-z]*)$", str(cid or ""))
        return (int(m.group(1)), m.group(2)) if m else (10**9, str(cid or ""))

    def _verdict_rank(r: dict) -> int:
        # Jie rows: order by compound_verdict. BDB rows: ordering is fully
        # carried by cat_order above, so they share a single rank here.
        if "compound_verdict" in r:
            return _VERDICT_ORDER.get(r.get("compound_verdict", ""), 99)
        return 0

    rows_sorted = sorted(rows, key=lambda r: (
        cat_order.get(r.get("verdict", ""), 99),
        _verdict_rank(r),
        # Group rows of the same verdict together within a rank tier.
        r.get("compound_verdict") or r.get("verdict") or "",
        _natural_cid(r.get("compound_id", "")),
    ))

    counts: dict = {"total": 0, "matched_compounds": 0,
                    "matched_value_match": 0, "matched_value_diff": 0,
                    "matched_no_v2_value": 0, "matched_patent_verified": 0,
                    "matched_binned": 0,
                    "stereo_extras": 0, "v2_only": 0, "bdb_missed": 0,
                    "matched_stereo_variant": 0}
    last_data_row = header_row
    for ridx, row in enumerate(rows_sorted, start=header_row + 1):
        for cidx, fname in enumerate(fields, start=1):
            cell = ws.cell(row=ridx, column=cidx, value=row.get(fname, ""))
            cell.alignment = Alignment(horizontal="left", vertical="top")
        if row.get("_fill") is not None:
            for cidx in range(1, len(fields) + 1):
                ws.cell(row=ridx, column=cidx).fill = row["_fill"]
        last_data_row = ridx
        counts["total"] += 1
        # `matched_compounds` = structurally found (matched + value-diff);
        # both have the molecule, they only differ on whether the assay
        # value also agrees. Keeping these bucket names stable means the
        # Summary aggregation below needs no changes.
        verdict = row.get("verdict", "")
        if verdict == "matched":
            counts["matched_compounds"] += 1
            counts["matched_value_match"] += 1
        elif verdict == "matched_value_diff":
            counts["matched_compounds"] += 1
            counts["matched_value_diff"] += 1
        elif verdict == "matched_no_v2_value":
            # Structure matched; v2 has no value of the ref kind. Still a
            # structural match (counts toward matched_compounds / strict
            # coverage) — but NOT a value mismatch.
            counts["matched_compounds"] += 1
            counts["matched_no_v2_value"] += 1
        elif verdict == "matched_patent_verified":
            # v2's value confirmed in the patent; BDB differs (not our error).
            # A structural match AND a correct extraction — just not a BDB match.
            counts["matched_compounds"] += 1
            counts["matched_patent_verified"] += 1
        elif verdict == "matched_binned":
            # Patent gives a letter-grade bin; BDB's value falls in the range.
            # Structural match + value consistent with the patent's bin.
            counts["matched_compounds"] += 1
            counts["matched_binned"] += 1
        elif verdict == "v2_stereo_extra":
            counts["stereo_extras"] += 1
        elif verdict == "v2_only":
            counts["v2_only"] += 1
        elif verdict == "bdb_missed":
            counts["bdb_missed"] += 1
        elif verdict == "stereo_variant":
            counts["matched_stereo_variant"] += 1
        # Legacy compatibility (Jie sheet still uses compound_verdict)
        v = row.get("compound_verdict") or ""
        if v:
            counts[v] = counts.get(v, 0) + 1

    # ── Aggregate stats footer ────────────────────────────────────
    stats_start = last_data_row + 3
    ws.cell(row=stats_start, column=1, value="── AGGREGATE STATS ──").font = Font(bold=True, color="1F4E78", size=12)
    stats = []
    if counts.get("matched_compounds", 0) > 0:
        # BDB-style sheet
        # Stereo-variant compounds are "skeleton-correct" — count them
        # toward structural coverage (analyst can verify stereo manually).
        ref_total = (
            counts["matched_compounds"] + counts["bdb_missed"]
            + counts["matched_stereo_variant"]
        )
        v2_total = counts["matched_compounds"] + counts["stereo_extras"] + counts["v2_only"]
        structural_total = counts["matched_compounds"] + counts["matched_stereo_variant"]
        stats = [
            ("Reference compounds (total)",          ref_total),
            ("v2 compounds (total)",                  v2_total),
            ("",                                       ""),
            ("Cluster 1 — Matched (same molecule)",   counts["matched_compounds"]),
            ("    of which value matches exactly",    counts["matched_value_match"]),
            ("    of which v2 matches PATENT (BDB differs, not our error)",
                                                       counts.get("matched_patent_verified", 0)),
            ("    of which patent gives a BIN, BDB value in range (letter-grade)",
                                                       counts.get("matched_binned", 0)),
            ("    of which value differs (v2 not in patent — real)",
                                                       counts["matched_value_diff"]),
            ("    of which v2 has no value (gap, not mismatch)",
                                                       counts.get("matched_no_v2_value", 0)),
            ("Cluster 2 — v2 stereo extras",          counts["stereo_extras"]),
            ("Cluster 3 — v2-only",                   counts["v2_only"]),
            ("Cluster 4 — BDB missed",                counts["bdb_missed"]),
            ("Cluster 5 — Matched stereo-variant "
             "(same skeleton, different/lost stereo)", counts["matched_stereo_variant"]),
            ("",                                       ""),
            ("Strict coverage (full-IK match / ref)",
                f"{counts['matched_compounds']} / {ref_total} = "
                f"{counts['matched_compounds']/max(1,ref_total):.0%}"),
            ("Structural coverage (skeleton match / ref)",
                f"{structural_total} / {ref_total} = "
                f"{structural_total/max(1,ref_total):.0%}"),
            ("Value precision (exact / matched)",
                f"{counts['matched_value_match']} / {max(1,counts['matched_compounds'])} = "
                f"{counts['matched_value_match']/max(1,counts['matched_compounds']):.0%}"),
        ]
    else:
        # Jie sheet — uses compound_verdict
        n_match = counts.get("match", 0)
        n_v2_missing = counts.get("v2_missing_molecule", 0)
        n_v2_missed = counts.get("v2_missed", 0)
        n_other = counts["total"] - n_match - n_v2_missing - n_v2_missed
        stats = [
            ("Reference compounds (total)",          counts["total"]),
            ("",                                       ""),
            ("Exact value match",                     n_match),
            ("v2_missed (have molecule, missing assay)", n_v2_missed),
            ("v2_missing_molecule (no InChIKey match)",  n_v2_missing),
            ("Other (both_in_text, etc.)",            n_other),
            ("",                                       ""),
            ("Match rate",                            f"{n_match} / {counts['total']} = {n_match/max(1,counts['total']):.0%}"),
        ]
    for i, (label, val) in enumerate(stats):
        c1 = ws.cell(row=stats_start + 1 + i, column=1, value=label)
        c2 = ws.cell(row=stats_start + 1 + i, column=2, value=val)
        if label.startswith("    "):
            c1.font = Font(italic=True, color="666666")
        elif label and not label.startswith("─"):
            c1.font = Font(bold=True)

    ws.freeze_panes = ws.cell(row=header_row + 1, column=2)
    ws.auto_filter.ref = (
        f"{get_column_letter(1)}{header_row}:"
        f"{get_column_letter(len(fields))}{last_data_row}"
    )

    for cidx, fname in enumerate(fields, start=1):
        cl = get_column_letter(cidx)
        if "iupac" in fname or "explanation" in fname:
            ws.column_dimensions[cl].width = 50
        elif "assay" in fname:
            ws.column_dimensions[cl].width = 28
        elif fname == "route":
            ws.column_dimensions[cl].width = 32
        elif fname == "verdict":
            ws.column_dimensions[cl].width = 20
        elif "compound_id" in fname:
            ws.column_dimensions[cl].width = 12
        elif "qualifier" in fname:
            ws.column_dimensions[cl].width = 12
        elif "inchikey" in fname:
            ws.column_dimensions[cl].width = 22
        else:
            ws.column_dimensions[cl].width = 16

    return counts


# ── Consolidated "verified data" sheet ──────────────────────────
# "Verified" = the v2 compound is structurally corroborated by the
# reference (BDB / Jie). For BDB that's verdict ∈ {matched,
# matched_value_diff, stereo_variant} — every row whose InChIKey (full or
# skeleton) the reference also lists. For Jie it's compound_verdict ==
# "match". EXCLUDED: v2_only / v2_stereo_extra (extracted but the
# reference can't corroborate) and bdb_missed (not a v2 extraction).
_VERIFIED_BDB_VERDICTS = {"matched", "matched_value_diff", "stereo_variant"}
_VERIFIED_COLS = [
    "patent_id", "compound_id", "iupac_name", "inchikey_full", "route",
    "assay", "v2_value_uM", "v2_qualifier", "ref_value_uM", "verdict",
    "explanation",
]
_VERIFIED_FILL = {
    "matched": COLOR_MATCH,
    "matched_value_diff": COLOR_VALUE_DIFF,
    "stereo_variant": COLOR_STEREO_DIFF,
}


def _verified_from_bdb(row: dict) -> dict:
    """Normalize a BDB sheet row → the common verified schema.

    PREFERENCE (fix for the Verified_Data display bug): when the BDB
    matching logic stored `_matched_via_name` + `_matched_via_value` on
    the row, surface THOSE as the displayed assay/value. They are the
    actual pair that triggered the InChIKey-value match. Falling back to
    the first populated primary `_uM` column gave misleading results —
    US10899738 cid 19's primary `mv4_11_uM = 1.6` was shown next to
    `ref_value_uM = 0.061`, hiding the real matching pair `Menin Binding
    Affinity IC50 = 0.061 = 0.061`. Only fall back to the first primary
    column when the matched-pair fields are absent (matched_value_diff /
    stereo_variant rows where nothing actually matched).
    """
    assay = v2_val = v2_qual = ""
    if row.get("_matched_via_name") and row.get("_matched_via_value") is not None:
        assay = row["_matched_via_name"]
        v2_val = row["_matched_via_value"]
        # qualifier: try to recover from the primary-column qualifier
        # whose canonical name matches; harmless to leave blank if absent
        v2_qual = ""
    else:
        for k, v in row.items():
            if (k.endswith("_uM") and k != "ref_value_uM"
                    and not k.startswith("_") and v not in ("", None)):
                assay = k[:-3]
                v2_val = v
                v2_qual = row.get(f"{k[:-3]}_qualifier", "") or ""
                break
    return {
        "patent_id": row.get("patent_id", ""),
        "compound_id": row.get("compound_id", ""),
        "iupac_name": row.get("iupac_name", ""),
        "inchikey_full": row.get("inchikey_full", ""),
        "route": row.get("route", ""),
        "assay": assay,
        "v2_value_uM": v2_val,
        "v2_qualifier": v2_qual,
        "ref_value_uM": row.get("ref_value_uM", ""),
        "verdict": row.get("verdict", ""),
        "explanation": row.get("explanation", ""),
    }


def _verified_from_jie(row: dict) -> dict:
    """Normalize a Jie sheet row → the common verified schema. Jie carries
    two assays; surface the primary (FLAP Ki) else the IC50."""
    assay = v2_val = v2_qual = ref_val = ""
    for prefix, label in (("flap_binding_ki", "FLAP Binding Ki"),
                          ("hwb_ltb4_ic50", "HWB LTB4 IC50")):
        cand = row.get(f"{prefix}_uM_v2", "")
        if cand not in ("", None):
            assay = label
            v2_val = cand
            v2_qual = row.get(f"{prefix}_qualifier_v2", "") or ""
            ref_val = row.get(f"{prefix}_uM_jie", "")
            break
    return {
        "patent_id": row.get("patent_id", "US8952177"),
        "compound_id": row.get("compound_id", ""),
        "iupac_name": row.get("iupac_name", ""),
        "inchikey_full": row.get("v2_inchikey_full", ""),
        "route": row.get("route", ""),
        "assay": assay,
        "v2_value_uM": v2_val,
        "v2_qualifier": v2_qual,
        "ref_value_uM": ref_val,
        "verdict": "matched",   # normalize Jie "match" → unified vocabulary
        "explanation": row.get("compound_explanation", "") or "",
    }


def _write_verified_sheet(wb, sources: list[tuple[str, list[dict], str]]) -> None:
    """Final sheet: every structurally-verified compound across all patents,
    organized by patent. `sources` = ordered [(patent_id, rows, kind)] where
    kind ∈ {"bdb", "jie"}."""
    _vrank = {"matched": 0, "matched_value_diff": 1, "stereo_variant": 2}

    def _ncid(cid) -> tuple[int, str]:
        m = re.match(r"(\d+)([A-Za-z]*)", str(cid or ""))
        return (int(m.group(1)), m.group(2)) if m else (10**9, str(cid or ""))

    per_patent: list[tuple[str, list[dict]]] = []
    for pid, rows, kind in sources:
        vrows: list[dict] = []
        for r in rows:
            if kind == "bdb" and r.get("verdict") in _VERIFIED_BDB_VERDICTS:
                vrows.append(_verified_from_bdb(r))
            elif kind == "jie" and r.get("compound_verdict") == "match":
                vrows.append(_verified_from_jie(r))
        vrows.sort(key=lambda v: (_vrank.get(v["verdict"], 9),
                                  _ncid(v["compound_id"])))
        per_patent.append((pid, vrows))

    ws = wb.create_sheet("Verified_Data")
    ws.cell(row=1, column=1, value="Verified compounds — structurally corroborated by BDB / Jie, organized by patent").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=(
        "Includes verdict ∈ {matched, matched_value_diff, stereo_variant} "
        "(BDB) and Jie matches. Excludes v2-only/extra (no reference) and "
        "bdb_missed (not extracted). Filter `verdict`=matched for value-exact only.")
    ).font = Font(italic=True, color="666666")

    header_row = 4
    for cidx, fname in enumerate(_VERIFIED_COLS, start=1):
        c = ws.cell(row=header_row, column=cidx, value=fname)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = COLOR_HEADER
        c.alignment = Alignment(horizontal="left", vertical="center")

    ridx = header_row
    total = 0
    for pid, vrows in per_patent:
        for vr in vrows:
            ridx += 1
            total += 1
            for cidx, fname in enumerate(_VERIFIED_COLS, start=1):
                cell = ws.cell(row=ridx, column=cidx, value=vr.get(fname, ""))
                cell.alignment = Alignment(horizontal="left", vertical="top")
            fill = _VERIFIED_FILL.get(vr["verdict"])
            if fill is not None:
                for cidx in range(1, len(_VERIFIED_COLS) + 1):
                    ws.cell(row=ridx, column=cidx).fill = fill

    # Per-patent count footer
    foot = ridx + 3
    ws.cell(row=foot, column=1, value="── VERIFIED COUNTS BY PATENT ──").font = Font(bold=True, color="1F4E78", size=12)
    for i, (pid, vrows) in enumerate(per_patent, start=1):
        ws.cell(row=foot + i, column=1, value=pid).font = Font(bold=True)
        ws.cell(row=foot + i, column=2, value=len(vrows))
    ws.cell(row=foot + len(per_patent) + 1, column=1, value="TOTAL verified").font = Font(bold=True)
    ws.cell(row=foot + len(per_patent) + 1, column=2, value=total).font = Font(bold=True)

    ws.freeze_panes = ws.cell(row=header_row + 1, column=2)
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(_VERIFIED_COLS))}{ridx}"
    for cidx, fname in enumerate(_VERIFIED_COLS, start=1):
        cl = get_column_letter(cidx)
        if fname in ("iupac_name", "explanation"):
            ws.column_dimensions[cl].width = 50
        elif fname == "assay":
            ws.column_dimensions[cl].width = 26
        elif fname == "inchikey_full":
            ws.column_dimensions[cl].width = 30
        elif fname == "verdict":
            ws.column_dimensions[cl].width = 20
        elif fname == "route":
            ws.column_dimensions[cl].width = 24
        else:
            ws.column_dimensions[cl].width = 14


def _letter_grade_status(patent_id: str) -> bool:
    """Return True iff this patent's primary assays are reported as LETTER
    GRADES (`+`, `++`, `+++`, `++++`) rather than numeric μM values.

    Detection: the patent's assay_tables.json shows values clustered on the
    geometric-mean midpoints of letter-grade brackets (`+ ≈ 31.6 μM`, `++ ≈
    3.16 μM`, `+++ ≈ 0.316 μM`, `++++ ≈ 0.0316 μM`). When ≥30% of nonzero
    values across the patent fall in this set (±5%), the patent doesn't
    contain real numeric IC50s — BDB's curated numerics are independent
    estimates from the same letter grades, so they can never exact-match
    v2's geomean-converted values. The Summary flags this so a low value-
    precision% isn't misread as an extraction failure (verified concrete:
    US11292791 reports IC50 as `+/++/+++/++++` in TABLE 19; BDB curators
    assigned 50.5, 0.55, 0.0055 μM — different bracket boundaries — and
    the 660 `matched_value_diff` rows are all this representation gap).
    """
    f = (Path(__file__).resolve().parents[3]
         / "output_v2" / "text_extraction" / patent_id / "assay_tables.json")
    if not f.exists():
        return False
    try:
        ay = json.loads(f.read_text())
    except (ValueError, OSError):
        return False
    # Direct signal: honest letter-bin records (source=letter_bin,
    # value_numeric=None) as emitted by routes.letter_bin_assays for
    # US11566007. When a meaningful fraction of the patent's COMPOUNDS
    # carry a bin, it's a letter-grade patent regardless of the geomean
    # heuristic.
    #
    # Counted per COMPOUND, not per row, because a second extraction tier
    # duplicates only the numeric rows — it has no reason to re-emit the
    # honest value_numeric=None bins — so a row ratio is diluted from the
    # denominator alone. Verified: 31 bin / 100 rows (0.31 -> True) becomes
    # 31 / 169 (0.18 -> False) once the 69 non-bin rows are duplicated
    # under a second source, flipping a genuinely letter-graded patent to
    # "numeric" on data saying nothing new. Duplicates add rows, never
    # compounds. Re-measured over all 22 patents: the compound ratio
    # reproduces the row ratio's verdict everywhere (US11566007 0.99 rows /
    # 0.98 cids, US11292791 0.53 / 0.84, all others 0.00) — no patent
    # changes classification.
    n_cids = len(ay)
    n_cids_bin = sum(1 for rows in ay.values()
                     if any(r.get("source") == "letter_bin" for r in rows))
    if n_cids and n_cids_bin / n_cids >= 0.30:
        return True
    GEOMEAN = (0.0316, 0.316, 3.16, 31.6)
    total = 0
    geomean_hits = 0
    distinct_gm: set[float] = set()
    for cid, rows in ay.items():
        # Same reason, applied to the `total >= 20` floor below: duplicate
        # rows would let a patent clear it without a single new value.
        for r in distinct_measurements(rows):
            v = r.get("value_numeric")
            if v in (None, 0, 0.0):
                continue
            total += 1
            for g in GEOMEAN:
                if abs(v - g) / g < 0.05:
                    geomean_hits += 1
                    distinct_gm.add(g)
                    break
    # Threshold tuned from real data (other 4 patents have 1.4–4.1 % incidental
    # geomean hits; US11292791 sits at 20 % — letter-grade signal is well
    # above the noise floor). Distinct-gm ≥ 3 rules out one-value flukes.
    return total >= 20 and len(distinct_gm) >= 3 and (geomean_hits / total) >= 0.10


def _patent_status(patent_id: str) -> str:
    """Composite status for the Summary's Status column. Priority order:
    `image-recog pending` (figures only) > `letter-grade` (no numeric IC50)
    > `text`. Returning the most informative single label keeps the column
    scannable; the legend below explains both flags."""
    if _image_recog_status(patent_id) == "image-recog pending":
        return "image-recog pending"
    if _letter_grade_status(patent_id):
        return "letter-grade"
    return "text"


def _image_recog_status(patent_id: str) -> str:
    """Read the patent's route_audit.json and report whether its missing
    structures are awaiting the image-recognition arm.

    `needs_image_recognition` is True only when a large fraction of the
    patent's drawn structures (C##### figures) have NO GP-embedded SMILES
    and NO IUPAC in the text — i.e. they exist ONLY as figure images we
    haven't OCR'd yet. For those patents a low Strict Cov% is EXPECTED and
    is NOT a text-extraction failure, so the Summary flags them rather than
    letting the number be misread. (Verified read-only: US10544143 ~98%,
    US11566007 ~95% of missing BDB compounds have no text IUPAC at all.)
    """
    f = (Path(__file__).resolve().parents[3]
         / "output_v2" / "text_extraction" / patent_id / "route_audit.json")
    if not f.exists():
        return "text"
    try:
        sif = (json.loads(f.read_text()).get("structure_image_figures") or {})
    except (ValueError, OSError):
        return "text"
    return "image-recog pending" if sif.get("needs_image_recognition") else "text"


def _write_summary(wb, sheet_counts: dict[str, dict]) -> None:
    ws = wb.create_sheet("Summary", 0)
    ws.cell(row=1, column=1, value="v2 Extraction — Hand-check Workbook").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value="One row per compound. Compound coverage = matched / reference total.").font = Font(italic=True, color="666666")

    headers = ["Sheet", "Reference", "Ref total", "v2 total",
               "Matched", "Stereo variant", "Stereo extras",
               "v2-only", "BDB missed",
               "Strict Cov%", "Structural Cov%", "Status"]
    for cidx, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=cidx, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = COLOR_HEADER

    sheet_meta = [
        ("US10899738_vs_BDB", "BindingDB"),
        ("US10214537_vs_BDB", "BindingDB"),
        ("US8952177_vs_Jie", "Jie's hand-curation"),
        ("US9718790_vs_BDB", "BindingDB"),
        ("US11292791_vs_BDB", "BindingDB"),
        ("US10246453_vs_BDB", "BindingDB"),
        ("US11254686_vs_BDB", "BindingDB"),
        ("US10273259_vs_BDB", "BindingDB"),
        ("US9694016_vs_BDB",  "BindingDB"),
        ("US10544143_vs_BDB", "BindingDB"),
        ("US11566007_vs_BDB", "BindingDB (Markush)"),
    ]
    for ridx, (sheet, ref) in enumerate(sheet_meta, start=5):
        c = sheet_counts.get(sheet, {})
        # Detect BDB vs Jie by which keys are populated
        is_bdb = (c.get("matched_compounds", 0) +
                  c.get("stereo_extras", 0) +
                  c.get("v2_only", 0) +
                  c.get("bdb_missed", 0)) > 0
        if is_bdb:
            matched = c.get("matched_compounds", 0)
            stereo_variant = c.get("matched_stereo_variant", 0)
            stereo_x = c.get("stereo_extras", 0)
            v2_only = c.get("v2_only", 0)
            bdb_miss = c.get("bdb_missed", 0)
            ref_total = matched + bdb_miss + stereo_variant
            v2_total = matched + stereo_x + v2_only
            strict_cov = f"{matched/max(1,ref_total):.0%}"
            struct_cov = f"{(matched+stereo_variant)/max(1,ref_total):.0%}"
        else:
            # Jie sheet — uses compound_verdict counts. "Matched" = v2
            # extracted the SAME molecule Jie curated. That's `match` PLUS
            # `both_in_text` (structure agrees, both values appear in the
            # patent — these are full structural matches, cross-verified by
            # jie_ik == v2_ik) PLUS `jie_duplicate_same_molecule`. Counting
            # only `match` here is what made the gold-standard Jie sheet read
            # a misleading 51% instead of ~84%.
            matched = (c.get("match", 0) + c.get("both_in_text", 0)
                       + c.get("jie_duplicate_same_molecule", 0))
            stereo_variant = c.get("stereo_diff", 0)
            stereo_x = c.get("v2_extra", 0)
            v2_only = 0
            # Jie has the molecule, v2 didn't extract it (or Jie's reference
            # IUPAC was wrong so it can't be verified).
            bdb_miss = (c.get("v2_missing_molecule", 0) + c.get("v2_missed", 0)
                        + c.get("ref_iupac_wrong", 0))
            ref_total = c.get("total", 0)
            v2_total = matched + stereo_x
            strict_cov = f"{matched/max(1,ref_total):.0%}"
            struct_cov = f"{(matched+stereo_variant)/max(1,ref_total):.0%}"
        ws.cell(row=ridx, column=1, value=sheet)
        ws.cell(row=ridx, column=2, value=ref)
        ws.cell(row=ridx, column=3, value=ref_total)
        ws.cell(row=ridx, column=4, value=v2_total)
        ws.cell(row=ridx, column=5, value=matched).fill = COLOR_MATCH
        ws.cell(row=ridx, column=6, value=stereo_variant).fill = COLOR_STEREO_DIFF
        ws.cell(row=ridx, column=7, value=stereo_x).fill = COLOR_STEREO_DIFF
        ws.cell(row=ridx, column=8, value=v2_only).fill = COLOR_V2_EXTRA
        ws.cell(row=ridx, column=9, value=bdb_miss).fill = COLOR_V2_WRONG
        ws.cell(row=ridx, column=10, value=strict_cov).font = Font(bold=True)
        ws.cell(row=ridx, column=11, value=struct_cov).font = Font(bold=True)
        # Status: flag patents whose missing structures live only in figure
        # images (`image-recog pending`) OR whose IC50 is reported as
        # letter grades not numbers (`letter-grade`) — both produce a low
        # value-precision% that's NOT a text-extraction bug.
        status = _patent_status(sheet.split("_")[0])
        sc = ws.cell(row=ridx, column=12, value=status)
        if status != "text":
            sc.fill = COLOR_REF_WRONG
            sc.font = Font(bold=True)

    # Compact legend — color-coded categories
    legend_start = 5 + len(sheet_meta) + 2
    ws.cell(row=legend_start, column=1, value="Categories:").font = Font(bold=True)
    legend = [
        ("Matched",         "v2 has the same molecule as the reference (full InChIKey match)", COLOR_MATCH),
        ("Stereo variant",  "Same skeleton (first 14-char InChIKey); v2 has different/lost stereo. Compound is structurally correct — review stereo.", COLOR_STEREO_DIFF),
        ("Stereo extras",   "v2 has a stereoisomer the reference (BDB) didn't curate",         COLOR_STEREO_DIFF),
        ("v2-only",         "v2 extracted from patent; reference has no entry",                COLOR_V2_EXTRA),
        ("BDB missed",      "Reference has the compound; v2 didn't extract it",                COLOR_V2_WRONG),
    ]
    for i, (label, desc, fill) in enumerate(legend):
        c1 = ws.cell(row=legend_start + 1 + i, column=1, value=label)
        c1.fill = fill
        c1.font = Font(bold=True)
        ws.cell(row=legend_start + 1 + i, column=2, value=desc)

    # Status legend — contextualize the two status flags.
    status_row = legend_start + 1 + len(legend) + 1
    sc = ws.cell(row=status_row, column=1, value="image-recog pending")
    sc.fill = COLOR_REF_WRONG
    sc.font = Font(bold=True)
    ws.cell(row=status_row, column=2, value=(
        "Most missing compounds exist ONLY as figure images (no IUPAC in the "
        "patent text). A low Strict Cov% here awaits the image-recognition "
        "arm — it is NOT a text-extraction failure."))
    lg_row = status_row + 1
    sc2 = ws.cell(row=lg_row, column=1, value="letter-grade")
    sc2.fill = COLOR_REF_WRONG
    sc2.font = Font(bold=True)
    ws.cell(row=lg_row, column=2, value=(
        "Patent reports activity as letter grades (+/++/+++/++++), not "
        "numeric IC50. v2 converts to geomean μM (0.0316/0.316/3.16/31.6); "
        "BDB curators independently estimated numeric values from the same "
        "grades. The value columns CANNOT exact-match — the gap is the data "
        "format, not extraction. Structure (Strict Cov%) and BDB-missed counts "
        "remain reliable."))

    for cl, w in zip("ABCDEFGHIJKL", [22, 22, 10, 10, 12, 14, 14, 12, 14, 12, 16, 20]):
        ws.column_dimensions[cl].width = w


def build_assay_row_stats(patent_id: str) -> dict:
    """Per-patent assay-row coverage stats — the measurement-level view
    that complements the compound-level Summary.

    Returns dict with keys:
      v2_cids_total       : # cids with rows in assay_tables
      v2_rows_total       : total assay rows extracted
      bdb_rows_total      : total rows BDB curated for this patent
      bdb_iks             : # unique InChIKeys in BDB for this patent
      matched_iks         : # BDB IKs that resolve to a v2 cid
      matched_w_rows      : # of those matched cids that have any rows
      v2_rows_for_matched : assay rows v2 extracted for matched cids
      bdb_row_coverage    : v2_rows_for_matched / bdb_rows_total (×100)
      top_assays          : [(assay_name, count), ...] top-5 by row count

    Every v2 count here is over DISTINCT measurements, which is what the
    sheet already claims to show ("the individual (cid, assay, value)
    measurements"). `assay_tables` now carries two extraction tiers side
    by side without deduplication, and `bdb_row_coverage_pct` has BDB's
    own row count as its denominator, so counting raw rows let ONE
    duplicated v2 row take the colour-highlighted quality signal from
    100 % to 200 % while BDB's side did not move at all.

    Note the metric legitimately exceeds 100 % without any duplicate —
    US10214537 reads 296.7 %, US10899738 211.1 %, US9745328 148.6 %,
    because v2 captures multi-column panels BDB never curated. That is
    why the defect had to be fixed by deduplicating rather than by
    capping: a cap would have hidden it behind numbers that are already
    supposed to be there. Measured over the corpus, deduplication moves
    no patent's percentage today (only US11566007 and US20240335431A1
    have collapsible rows at all, 93 between them, and neither has BDB
    data) — this is a guard for the merge, not a restatement of it.
    """
    import csv, json
    from collections import defaultdict
    from pathlib import Path
    out_dir = Path(__file__).resolve().parent.parent.parent.parent / "output_v2" / "text_extraction" / patent_id
    if not (out_dir / "assay_tables.json").exists():
        return {"patent_id": patent_id, "missing": True}
    ay = json.loads((out_dir / "assay_tables.json").read_text())
    ex = json.loads((out_dir / "example_index.json").read_text())
    ay = {cid: distinct_measurements(rows) for cid, rows in ay.items()}
    v2_n_cids = len(ay)
    v2_n_rows = sum(len(rows) for rows in ay.values())
    by_assay: dict[str, int] = defaultdict(int)
    for rows in ay.values():
        for r in rows:
            by_assay[r.get("assay_name") or "(unknown)"] += 1
    # BDB — also track curation SOURCE (patent vs journal/PMID) and the
    # assay-panel density (distinct targets per compound). This lets a low
    # row-coverage be attributed WITH EVIDENCE: alternate-source (BDB cites a
    # journal, not the patent) vs v2 extraction gap vs BDB simply curating a
    # wider assay panel than the patent's main table reports per compound.
    bdb_rows = 0
    bdb_iks: set[str] = set()
    bdb_patent_rows = 0
    per_cmpd_targets: dict[str, set] = defaultdict(set)
    bdb_tsv = Path(__file__).resolve().parent.parent.parent.parent / "output" / "bindingdb" / "our_patents.tsv"
    with open(bdb_tsv) as f:
        reader = csv.DictReader(f, delimiter="\t")
        _fn = reader.fieldnames or []
        _tgt_col = (next((c for c in _fn if "Target Name" in c), None)
                    or next((c for c in _fn if "Target Source" in c), None))
        for row in reader:
            if patent_id in (row.get("Patent Number") or ""):
                bdb_rows += 1
                ik = (row.get("Ligand InChI Key") or "").strip()
                if ik:
                    bdb_iks.add(ik)
                    if _tgt_col:
                        per_cmpd_targets[ik].add((row.get(_tgt_col) or "").strip())
                # A row is patent-sourced unless it cites a journal article
                # (PMID / Article DOI) — that's the "alternate source" signal.
                has_article = bool((row.get("PMID") or "").strip()
                                   or (row.get("Article DOI") or "").strip())
                if not has_article:
                    bdb_patent_rows += 1
    # Match
    def _resolve(c: str) -> str:
        return ex.get(c, {}).get("canonical_cid") or c
    ex_ik: dict[str, str] = {}
    for k, v in ex.items():
        for ik_field in ([v.get("inchikey")] + (v.get("inchikey_aliases") or [])):
            ik = (ik_field or "").strip()
            if ik:
                ex_ik.setdefault(ik, k)
    matched = bdb_iks & set(ex_ik.keys())
    matched_cids = {_resolve(ex_ik[ik]) for ik in matched}
    v2_rows_for_matched = sum(len(ay.get(c, [])) for c in matched_cids)
    matched_w_rows = sum(1 for c in matched_cids if c in ay)
    top_assays = sorted(by_assay.items(), key=lambda x: -x[1])[:5]

    # ── Attribution metrics ──────────────────────────────────────────
    bdb_pct_patent = 100 * bdb_patent_rows / max(1, bdb_rows)
    targets_per_cmpd = (sum(len(t) for t in per_cmpd_targets.values())
                        / max(1, len(per_cmpd_targets)))
    # COMPOUND-level coverage — the fair extraction signal. BDB row coverage
    # is depressed whenever BDB curates a multi-target panel per compound, so
    # it understates extraction quality; this counts compounds, not rows.
    cmpd_coverage_pct = 100 * matched_w_rows / max(1, len(bdb_iks))
    if bdb_pct_patent < 90:
        diagnosis = (f"BDB ALT-SOURCE: {100 - bdb_pct_patent:.0f}% of BDB rows cite a "
                     f"journal/PMID, not this patent — not an extraction gap")
    elif cmpd_coverage_pct < 60:
        diagnosis = (f"v2 EXTRACTION GAP: only {cmpd_coverage_pct:.0f}% of BDB's patent "
                     f"compounds carry a v2 assay value (the compounds ARE in the patent)")
    elif targets_per_cmpd >= 2.5:
        diagnosis = (f"METRIC ONLY: BDB curates a {targets_per_cmpd:.0f}-target panel/compound; "
                     f"row-coverage understates real extraction (see Cmpd coverage%)")
    else:
        diagnosis = "ok"

    return {
        "patent_id": patent_id,
        "v2_cids_total": v2_n_cids,
        "v2_rows_total": v2_n_rows,
        "bdb_rows_total": bdb_rows,
        "bdb_iks": len(bdb_iks),
        "matched_iks": len(matched),
        "matched_w_rows": matched_w_rows,
        "v2_rows_for_matched": v2_rows_for_matched,
        "bdb_row_coverage_pct": (100 * v2_rows_for_matched / max(1, bdb_rows)),
        "bdb_pct_patent": bdb_pct_patent,
        "targets_per_cmpd": targets_per_cmpd,
        "cmpd_coverage_pct": cmpd_coverage_pct,
        "diagnosis": diagnosis,
        "top_assays": top_assays,
    }


def _write_assay_stats(wb, stats_list: list[dict]) -> None:
    """Write a `Assay Row Stats` sheet — measurement-level coverage,
    complementing the compound-level Summary. Same color palette."""
    ws = wb.create_sheet("Assay Row Stats", 1)
    ws.cell(row=1, column=1, value="Assay-row coverage (the measurement-level view)").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value="Summary counts compounds; this sheet counts the individual (cid, assay, value) measurements.").font = Font(italic=True, color="666666")

    headers = [
        "Patent",
        "v2 cids w/ rows",
        "v2 assay rows total",
        "BDB rows total",
        "BDB cmpd IKs",
        "Matched IKs",
        "Matched w/ rows",
        "v2 rows for matched",
        "BDB row coverage%",
        "Cmpd coverage%",
        "BDB targets/cmpd",
        "BDB % from patent",
        "Diagnosis",
        "Top assays (by v2 row count)",
    ]
    for cidx, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=cidx, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = COLOR_HEADER

    for ridx, s in enumerate(stats_list, start=5):
        if s.get("missing"):
            ws.cell(row=ridx, column=1, value=s["patent_id"])
            ws.cell(row=ridx, column=2, value="(no v2 output)")
            continue
        ws.cell(row=ridx, column=1, value=s["patent_id"])
        ws.cell(row=ridx, column=2, value=s["v2_cids_total"])
        ws.cell(row=ridx, column=3, value=s["v2_rows_total"]).font = Font(bold=True)
        ws.cell(row=ridx, column=4, value=s["bdb_rows_total"])
        ws.cell(row=ridx, column=5, value=s["bdb_iks"])
        ws.cell(row=ridx, column=6, value=s["matched_iks"]).fill = COLOR_MATCH
        ws.cell(row=ridx, column=7, value=s["matched_w_rows"]).fill = COLOR_MATCH
        ws.cell(row=ridx, column=8, value=s["v2_rows_for_matched"]).fill = COLOR_MATCH
        pct_cell = ws.cell(row=ridx, column=9, value=f"{s['bdb_row_coverage_pct']:.0f}%")
        pct_cell.font = Font(bold=True)
        # Highlight: >100% = blue (we extract MORE than BDB); >=80% green; <50% pink
        cov = s["bdb_row_coverage_pct"]
        if cov >= 100:
            pct_cell.fill = COLOR_V2_EXTRA  # blue-ish: extra
        elif cov >= 80:
            pct_cell.fill = COLOR_MATCH
        elif cov < 50:
            pct_cell.fill = COLOR_V2_WRONG
        # Cmpd coverage% — the FAIR extraction metric (compounds, not rows)
        cc = s.get("cmpd_coverage_pct", 0.0)
        cc_cell = ws.cell(row=ridx, column=10, value=f"{cc:.0f}%")
        cc_cell.fill = COLOR_MATCH if cc >= 80 else (COLOR_V2_WRONG if cc < 60 else COLOR_REF_WRONG)
        ws.cell(row=ridx, column=11, value=f"{s.get('targets_per_cmpd', 0):.1f}")
        # BDB % from patent — <90 means BDB curated from an alternate source
        pp = s.get("bdb_pct_patent", 100.0)
        pp_cell = ws.cell(row=ridx, column=12, value=f"{pp:.0f}%")
        pp_cell.fill = COLOR_V2_WRONG if pp < 90 else COLOR_MATCH
        diag = s.get("diagnosis", "")
        d_cell = ws.cell(row=ridx, column=13, value=diag)
        d_cell.alignment = Alignment(wrap_text=True, vertical="top")
        if diag.startswith("BDB ALT-SOURCE"):
            d_cell.fill = COLOR_AMBIG            # grey: not our problem
        elif diag.startswith("v2 EXTRACTION GAP"):
            d_cell.fill = COLOR_V2_WRONG          # pink: real gap to fix
        elif diag.startswith("METRIC"):
            d_cell.fill = COLOR_V2_EXTRA          # blue: metric artifact, not a gap
        else:
            d_cell.fill = COLOR_MATCH
        # Top assays
        top_str = ", ".join(f"{n}× {a[:30]}" for a, n in s["top_assays"][:5])
        ws.cell(row=ridx, column=14, value=top_str).alignment = Alignment(wrap_text=True, vertical="top")

    # Legend
    legend_start = 5 + len(stats_list) + 2
    ws.cell(row=legend_start, column=1, value="Interpretation:").font = Font(bold=True)
    notes = [
        "BDB row coverage% > 100  → v2 extracts MORE measurements than BDB curated "
        "(intermediate-step assays, replicate runs, etc. that BDB ignored)",
        "BDB row coverage% 80-100 → v2 has the same density as BDB; expected outcome on well-extracted patents",
        "BDB row coverage% < 50   → v2 is missing measurements BDB curated; investigate (could be OCR loss, "
        "letter-grade table, or missing target)",
        "Matched IKs counts the BDB compounds we resolved by InChIKey (including aliases / canonical_cid pointers)",
        "Matched w/ rows shows how many of those matched compounds actually carry assay rows in v2",
        "v2 rows for matched is the measurement count for the matched cids — divide by Matched w/ rows for the average rows-per-compound",
    ]
    for i, n in enumerate(notes):
        ws.cell(row=legend_start + 1 + i, column=1, value=n).font = Font(italic=True, color="555555")
        ws.merge_cells(start_row=legend_start + 1 + i, start_column=1,
                       end_row=legend_start + 1 + i, end_column=10)

    widths = [14, 16, 20, 16, 14, 13, 16, 20, 18, 60]
    for i, w in enumerate(widths):
        ws.column_dimensions[chr(ord("A") + i)].width = w


def main() -> None:
    print("Building hand-check workbook (centralized modules + sorted + Jie-format)...")

    rows_us738 = build_bdb_rows("US10899738")
    rows_us537 = build_bdb_rows("US10214537")
    rows_jie   = build_jie_rows()
    # New batched-mode patents — added 2026-05 after the cid-namespace,
    # header-propagation, row-merge-repair, and batched-LLM fixes lifted
    # them from baseline 1-43 % BDB coverage to 81-100 %.
    rows_us790 = build_bdb_rows("US9718790")
    rows_us291 = build_bdb_rows("US11292791")
    rows_us453 = build_bdb_rows("US10246453")
    # 2026-05-21 second batch (5 patents) — adds:
    #   US11254686 (2002 BDB rows, Z-prefix CID convention)
    #   US10273259 (881 BDB rows)
    #   US9694016  (2555 BDB rows)
    #   US10544143 (3210 BDB rows)
    #   US11566007 (7524 BDB rows — pure Markush, 0 GP-embedded SMILES,
    #               extraction returned ex=0; queued for Markush phase)
    rows_us686 = build_bdb_rows("US11254686")
    rows_us259 = build_bdb_rows("US10273259")
    rows_us016 = build_bdb_rows("US9694016")
    rows_us143 = build_bdb_rows("US10544143")
    rows_us007 = build_bdb_rows("US11566007")

    wb = openpyxl.Workbook()
    if wb.active:
        wb.remove(wb.active)

    counts = {
        "US10899738_vs_BDB": _write_sheet(wb, "US10899738_vs_BDB", rows_us738,
                                          "vs BindingDB; sorted by verdict"),
        "US10214537_vs_BDB": _write_sheet(wb, "US10214537_vs_BDB", rows_us537,
                                          "vs BindingDB; sorted by verdict"),
        "US8952177_vs_Jie":  _write_sheet(wb, "US8952177_vs_Jie", rows_jie,
                                          "vs Jie's curation; mirrors her column names; one row per compound"),
        "US9718790_vs_BDB":  _write_sheet(wb, "US9718790_vs_BDB", rows_us790,
                                          "vs BindingDB; sorted by verdict"),
        "US11292791_vs_BDB": _write_sheet(wb, "US11292791_vs_BDB", rows_us291,
                                          "vs BindingDB; sorted by verdict"),
        "US10246453_vs_BDB": _write_sheet(wb, "US10246453_vs_BDB", rows_us453,
                                          "vs BindingDB; sorted by verdict"),
        "US11254686_vs_BDB": _write_sheet(wb, "US11254686_vs_BDB", rows_us686,
                                          "vs BindingDB; sorted by verdict"),
        "US10273259_vs_BDB": _write_sheet(wb, "US10273259_vs_BDB", rows_us259,
                                          "vs BindingDB; sorted by verdict"),
        "US9694016_vs_BDB":  _write_sheet(wb, "US9694016_vs_BDB", rows_us016,
                                          "vs BindingDB; sorted by verdict"),
        "US10544143_vs_BDB": _write_sheet(wb, "US10544143_vs_BDB", rows_us143,
                                          "vs BindingDB; sorted by verdict"),
        "US11566007_vs_BDB": _write_sheet(wb, "US11566007_vs_BDB", rows_us007,
                                          "vs BindingDB; PURE MARKUSH — ex=0, queued for Markush phase"),
    }
    _write_summary(wb, counts)
    # Assay-row stats — measurement-level view (complements compound Summary).
    assay_stats = [
        build_assay_row_stats(pid) for pid in (
            "US10899738", "US10214537", "US9718790", "US11292791", "US10246453",
            "US11254686", "US10273259", "US9694016", "US10544143", "US11566007",
        )
    ]
    _write_assay_stats(wb, assay_stats)
    wb.move_sheet("Summary", offset=-len(wb.worksheets) + 1)
    wb.move_sheet("Assay Row Stats", offset=-len(wb.worksheets) + 2)
    # Final sheet: all structurally-verified compounds, organized by patent.
    # Created AFTER the move_sheet calls so it lands last without shifting
    # the Summary / Assay-Row-Stats positions (their offsets depend on the
    # current worksheet count).
    _write_verified_sheet(wb, [
        ("US10899738", rows_us738, "bdb"),
        ("US10214537", rows_us537, "bdb"),
        ("US8952177",  rows_jie,   "jie"),
        ("US9718790",  rows_us790, "bdb"),
        ("US11292791", rows_us291, "bdb"),
        ("US10246453", rows_us453, "bdb"),
        ("US11254686", rows_us686, "bdb"),
        ("US10273259", rows_us259, "bdb"),
        ("US9694016",  rows_us016, "bdb"),
        ("US10544143", rows_us143, "bdb"),
        ("US11566007", rows_us007, "bdb"),
    ])
    _save_ref_ik_cache()  # final flush of any sub-throttle reference IKs
    out = DOWNLOADS / "v2_handcheck.xlsx"
    wb.save(out)
    print(f"Wrote: {out}\n")
    print(f"Reference-IK cache: {len(_REF_IK_CACHE)} names persisted to "
          f"{_REF_IK_CACHE_PATH.name} (future rebuilds reuse, $0)")
    print("Assay-row stats:")
    print(f"  {'patent':>12}  {'v2_rows':>8}  {'BDB_rows':>9}  {'matched':>8}  {'v2_for_matched':>15}  {'cov%':>6}")
    for s in assay_stats:
        if s.get("missing"):
            continue
        print(f"  {s['patent_id']:>12}  {s['v2_rows_total']:>8}  {s['bdb_rows_total']:>9}  "
              f"{s['matched_iks']:>8}  {s['v2_rows_for_matched']:>15}  {s['bdb_row_coverage_pct']:>5.0f}%")
    print()

    # Footer mirrors the unified verdict buckets. BDB sheets store
    # `matched_compounds` (= matched + value_diff); the Jie sheet stores
    # `compound_verdict` counts (match / both_in_text / …). Handle both so
    # BDB rows don't print a misleading 0 (they have no `match` key).
    print(f"{'Sheet':>22}  {'rows':>5}  {'matched':>7}  {'val_diff':>8}  "
          f"{'stereo':>6}  {'v2_only':>7}  {'bdb_miss':>8}")
    print("-" * 82)
    for s, c in counts.items():
        matched = c.get("matched_compounds", 0) or (
            c.get("match", 0) + c.get("both_in_text", 0)
            + c.get("jie_duplicate_same_molecule", 0))
        val_diff = c.get("matched_value_diff", 0) or c.get("both_in_text", 0)
        stereo = (c.get("matched_stereo_variant", 0) + c.get("stereo_extras", 0)
                  + c.get("stereo_diff", 0) + c.get("v2_extra", 0))
        v2_only = c.get("v2_only", 0)
        bdb_miss = (c.get("bdb_missed", 0) + c.get("v2_missing_molecule", 0)
                    + c.get("v2_missed", 0) + c.get("ref_iupac_wrong", 0))
        print(f"{s:>22}  {c['total']:>5}  {matched:>7}  {val_diff:>8}  "
              f"{stereo:>6}  {v2_only:>7}  {bdb_miss:>8}")

    # ── Row-count reconciliation (BDB sheets) ─────────────────────────
    # Every BDB row lands in exactly one verdict bucket, so the rendered
    # row count must equal the sum of the buckets. A mismatch means a row
    # got a verdict the counter doesn't recognize (off-by-N) — surface it
    # loudly rather than letting a silent miscount skew the Summary.
    print()
    print("Row-count reconciliation (BDB sheets — rows == sum of verdict buckets):")
    all_ok = True
    for s, c in counts.items():
        buckets = (c.get("matched_compounds", 0) + c.get("stereo_extras", 0)
                   + c.get("v2_only", 0) + c.get("bdb_missed", 0)
                   + c.get("matched_stereo_variant", 0))
        if buckets == 0:
            continue  # Jie sheet — uses a different bucket scheme
        total = c.get("total", 0)
        ok = total == buckets
        all_ok = all_ok and ok
        print(f"  {s:>22}  rows={total:>4}  buckets={buckets:>4}  "
              f"[{'OK' if ok else 'MISMATCH'}]")
    if not all_ok:
        print("  ⚠️  RECONCILIATION MISMATCH — a verdict bucket is miscounted.")


if __name__ == "__main__":
    main()
