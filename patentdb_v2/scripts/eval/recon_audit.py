"""Reconciliation audit → spreadsheet (BLUEPRINT end-to-end, MinerU vs GP).

For each patent that has BOTH sources, runs the full deterministic pipeline
(adapters → measure → reconcile) and writes an OPENABLE xlsx so every choosing
decision is verifiable:
  • Summary    — per-patent resolution breakdown + agreement %
  • Conflicts  — the review queue: where the two sources disagreed (both values)
  • Reconciled — every chosen measurement with its resolution + confidence

Usage:  python3 -m patentdb.scripts.eval.recon_audit
Writes: ~/Downloads/reconciliation_audit.xlsx  (and opens it)
"""
from __future__ import annotations

import collections
import glob
import json
import os
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from patentdb.core.tables.gp_adapter import parse_gp_text
from patentdb.core.tables.measure import table_to_measurements
from patentdb.core.tables.mineru_adapter import parse_mineru_page
from patentdb.core.tables.reconcile import reconcile_with_llm
from patentdb.core.tables.resolution_memory import ResolutionMemory

REPO = Path(__file__).resolve().parents[3]
OUT = Path.home() / "Downloads" / "reconciliation_audit.xlsx"
PG = re.compile(r"page_(\d+)")
NAVY = PatternFill("solid", fgColor="1F4E78")
RED = PatternFill("solid", fgColor="F8CBAD")
GREEN = PatternFill("solid", fgColor="E2F0D9")
HDRF = Font(bold=True, color="FFFFFF")


def _mineru(pid):
    out, raw = [], []
    for pg in sorted(glob.glob(str(REPO / pid / "all_pages" / "page_*.md"))):
        txt = open(pg).read()
        raw.append(txt)
        mn = PG.search(pg)
        out += [m for t in parse_mineru_page(txt, page=int(mn.group(1)) if mn else 0)
                for m in table_to_measurements(t)]
    return out, "".join(raw)


def _gp(pid):
    p = REPO / "output_v2" / "gpatents_cache" / f"{pid}.json"
    raw = json.loads(p.read_text()).get("description", "") if p.exists() else ""
    return [m for t in parse_gp_text(pid, raw or None) for m in table_to_measurements(t)], raw


def _hdr(ws, cols, widths):
    ws.append(cols)
    for c, w in zip(ws[1], widths):
        c.fill, c.font = NAVY, HDRF
        ws.column_dimensions[c.column_letter].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"


def main() -> None:
    patents = sorted(os.path.basename(os.path.dirname(p))
                     for p in glob.glob(str(REPO / "*" / "all_pages")))
    mem = ResolutionMemory()
    summary, conflicts, reconciled = [], [], []
    for pid in patents:
        m, mraw = _mineru(pid)
        g, graw = _gp(pid)
        if not m and not g:
            continue
        rec = reconcile_with_llm(m, g, patent_id=pid, mineru_raw=mraw, gp_raw=graw,
                                 memory=mem)
        c = collections.Counter(r.provenance.get("resolution") for r in rec)
        resolved_x = (c.get("both_agree", 0) + c.get("conflict_resolved", 0)
                      + c.get("llm_reconciled", 0))
        comparable = resolved_x + c.get("conflict", 0)
        summary.append([pid, len(rec), c.get("both_agree", 0),
                        c.get("conflict_resolved", 0), c.get("llm_reconciled", 0),
                        c.get("conflict", 0), c.get("mineru_only", 0),
                        c.get("gp_only", 0),
                        f"{resolved_x / max(1, comparable):.0%}" if comparable else "—"])
        for r in rec:
            res = r.provenance.get("resolution")
            val = r.value_numeric if r.value_numeric is not None else r.value_raw
            reconciled.append([pid, r.cid, r.assay, val, r.unit, r.encoding, res,
                               r.provenance.get("confidence")])
            if res in ("conflict", "conflict_resolved", "llm_reconciled"):
                conflicts.append([pid, r.cid, r.assay, r.provenance.get("source"), val,
                                  r.provenance.get("other") or r.provenance.get("lost"),
                                  res, r.provenance.get("chose", "")])

    wb = Workbook()
    if wb.active is not None:
        wb.remove(wb.active)
    sm = wb.create_sheet("Summary")
    _hdr(sm, ["Patent", "Reconciled", "Both agree", "Det-resolved", "LLM-fixed",
              "Conflict (flagged)", "MinerU only", "GP only", "Resolved%"],
         [13, 11, 11, 13, 10, 16, 11, 9, 10])
    for r in summary:
        sm.append(r)
        if r[5]:                                  # flagged conflicts remain
            sm.cell(row=sm.max_row, column=6).fill = RED
        if r[4]:                                  # LLM-fixed present
            sm.cell(row=sm.max_row, column=5).fill = GREEN

    cf = wb.create_sheet("Conflicts (review)")
    _hdr(cf, ["Patent", "cid", "assay", "this source", "this value",
              "other value", "resolution", "chose"], [13, 8, 32, 11, 14, 14, 16, 8])
    for r in conflicts:
        cf.append(r)

    rc = wb.create_sheet("Reconciled")
    _hdr(rc, ["Patent", "cid", "assay", "value", "unit", "encoding",
              "resolution", "confidence"], [13, 9, 34, 12, 7, 9, 16, 10])
    for r in reconciled:
        rc.append(r)
        if r[6] == "both_agree":
            rc.cell(row=rc.max_row, column=7).fill = GREEN

    ms = wb.create_sheet("Resolution memory")
    _hdr(ms, ["key", "patent", "kind", "source_used", "confidence", "status",
              "#resolved", "reason"], [22, 14, 10, 11, 11, 9, 9, 42])
    for key, e in sorted(mem.d.items(), key=lambda kv: (kv[1].get("patent", ""), kv[0])):
        ms.append([key, e.get("patent", ""), e.get("kind", ""), e.get("source_used", ""),
                   e.get("confidence", ""), e.get("status", ""),
                   len(e.get("resolved", {})), e.get("reason", "")])
        if e.get("status") == "pending":
            ms.cell(row=ms.max_row, column=6).fill = RED

    wb.save(OUT)
    print(f"wrote {OUT}  ({len(reconciled)} measurements, {len(conflicts)} conflict rows, "
          f"{len(summary)} patents with both sources)")


if __name__ == "__main__":
    main()
