"""L1 audit → spreadsheet: prove the MinerU adapter captures tables faithfully.

One row per parsed table-block (filterable/sortable), so structural fidelity can
be checked without scrolling. Sheets:
  • Summary       — per-patent table counts
  • Assay tables  — only the assay-like blocks (the ones that matter)
  • All tables    — every parsed block

Usage:  python3 -m patentdb.scripts.eval.l1_audit
Writes: ~/Downloads/L1_table_audit.xlsx  (and opens it)
"""
from __future__ import annotations

import glob
import os
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from patentdb.core.tables.gp_adapter import parse_gp_text
from patentdb.core.tables.mineru_adapter import parse_mineru_page

REPO = Path(__file__).resolve().parents[3]
OUT = Path.home() / "Downloads" / "L1_table_audit.xlsx"
PG = re.compile(r"page_(\d+)")

# Preview-only assay signal (NOT the real L2 classifier) — just to filter the view.
_ASSAY_RE = re.compile(
    r"[IEAGC]\s?C\s?_?\s?50|K\s?[id]\b|CC\s?50|GI\s?50|pIC50|\bMIC\b|FRET|"
    r"binding|inhibit|HTRF|LTB|FLAP|antagonis|agonis|\bpEC|\bKi\b|enzymatic",
    re.I,
)
NAVY = PatternFill("solid", fgColor="1F4E78")
GREEN = PatternFill("solid", fgColor="E2F0D9")
HDRF = Font(bold=True, color="FFFFFF")

COLS = ["Patent", "Page", "Table", "Type", "Caption", "#cols", "#rows",
        "Header (column labels)", "Sample row 1", "Sample row 2", "Sample row 3"]
WIDTHS = [13, 6, 10, 7, 40, 6, 7, 52, 30, 30, 30]


def _is_assay(t) -> bool:
    return bool(_ASSAY_RE.search(t.caption + " " + " ".join(t.header)))


def _row(pid, t) -> list:
    samples = [" | ".join(r) for r in t.rows[:3]] + ["", "", ""]
    return [pid, t.page, t.table_label or "—", "assay" if _is_assay(t) else "other",
            t.caption[:300], t.n_cols, t.n_rows, " | ".join(t.header)[:400],
            *samples[:3]]


def _sheet(wb, name, rows, *, with_type=True):
    ws = wb.create_sheet(name)
    cols = COLS if with_type else [c for c in COLS if c != "Type"]
    widths = WIDTHS if with_type else [w for c, w in zip(COLS, WIDTHS) if c != "Type"]
    ws.append(cols)
    for c, w in zip(ws[1], widths):
        c.fill, c.font = NAVY, HDRF
        ws.column_dimensions[c.column_letter].width = w
    for r in rows:
        rr = r if with_type else [v for i, v in enumerate(r) if i != COLS.index("Type")]
        ws.append(rr)
        if (with_type and r[COLS.index("Type")] == "assay"):
            for c in ws[ws.max_row]:
                c.fill = GREEN
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{ws.max_row}"
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = Alignment(vertical="top", wrap_text=False)


def main() -> None:
    patents = sorted(
        os.path.basename(os.path.dirname(p))
        for p in glob.glob(str(REPO / "*" / "all_pages"))
    )
    all_rows, assay_rows, summary = [], [], []
    for pid in patents:
        tabs = []
        for pg in sorted(glob.glob(str(REPO / pid / "all_pages" / "page_*.md"))):
            mn = PG.search(pg)
            try:
                tabs.extend(parse_mineru_page(open(pg).read(),
                                              page=int(mn.group(1)) if mn else 0))
            except Exception:  # noqa: BLE001
                pass
        if not tabs:
            continue
        rows = [_row(pid, t) for t in tabs]
        a = [r for r in rows if r[COLS.index("Type")] == "assay"]
        all_rows.extend(rows)
        assay_rows.extend(a)
        summary.append([pid, len(rows), len(a), len(rows) - len(a)])

    wb = Workbook()
    if wb.active is not None:
        wb.remove(wb.active)               # drop the default empty sheet
    sm = wb.create_sheet("Summary")
    sm.append(["L1 MinerU table-structure audit — one row per parsed table block"])
    sm["A1"].font = Font(bold=True, size=13, color="1F4E78")
    sm.append([])
    sm.append(["Patent", "Tables", "Assay-like", "Other"])
    for c in sm[3]:
        c.fill, c.font = NAVY, HDRF
    for r in summary:
        sm.append(r)
    sm.append(["TOTAL", sum(r[1] for r in summary), sum(r[2] for r in summary),
               sum(r[3] for r in summary)])
    for c in sm[sm.max_row]:
        c.font = Font(bold=True)
    for col, w in zip("ABCD", [16, 9, 11, 8]):
        sm.column_dimensions[col].width = w
    sm.freeze_panes = "A4"

    _sheet(wb, "Assay tables (MinerU)", assay_rows, with_type=False)
    _sheet(wb, "All tables (MinerU)", all_rows, with_type=True)

    # GP-text candidate tables (the second source the reconciler will choose from)
    gp_rows = []
    for gp in sorted(glob.glob(str(REPO / "output_v2" / "gpatents_cache" / "*.json"))):
        pid = os.path.basename(gp)[:-5]
        try:
            ts = parse_gp_text(pid)
        except Exception:  # noqa: BLE001
            ts = []
        for t in ts:
            s = [" | ".join(r) for r in t.rows[:3]] + ["", "", ""]
            gp_rows.append([pid, t.n_cols, t.n_rows, " | ".join(t.header)[:400], *s[:3]])
    gw = wb.create_sheet("GP candidates")
    gcols = ["Patent", "#cols", "#rows", "Header (column labels)",
             "Sample row 1", "Sample row 2", "Sample row 3"]
    gw.append(gcols)
    for c, wdt in zip(gw[1], [13, 6, 7, 52, 30, 30, 30]):
        c.fill, c.font = NAVY, HDRF
        gw.column_dimensions[c.column_letter].width = wdt
    for r in gp_rows:
        gw.append(r)
    gw.freeze_panes = "A2"
    gw.auto_filter.ref = f"A1:{get_column_letter(len(gcols))}{gw.max_row}"

    wb.save(OUT)
    print(f"wrote {OUT}  (MinerU: {len(all_rows)} tables / {len(assay_rows)} assay; "
          f"GP: {len(gp_rows)} candidates; {len(summary)} patents)")


if __name__ == "__main__":
    main()
