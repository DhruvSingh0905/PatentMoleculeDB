"""Build the deliverable spreadsheet: N compounds, structure + every assay value.

    python3 -m patentdb.scripts.eval.build_jie_sample --n 3000

One row per compound in the wide sheet, laid out the way Jie asked for it —
`{assay}_{unit}`, `{assay}_qualifier`, `{assay}_n_runs`, and a `{assay}_assay`
column carrying the column header the patent actually printed. Every row is
attributed with `patent_id`.

Sampled from compounds that have BOTH a structure and at least one assay
value, because a row with no IUPAC is not usable for docking and a row with no
assay is not usable for triage. That intersection is smaller than either side:
structures come from `process_patent`, assays from the USPTO CALS path, and
only 20 patents have been through both.

Three sheets:
  compounds  — the wide format above, one row per compound
  assay_long — one row per (compound, assay). Lossless, and the sheet to pivot
               from. The wide sheet keeps only the first value per assay when a
               patent reports a compound twice; this one keeps both.
  README     — column dictionary, provenance, and what the numbers do not mean
"""
from __future__ import annotations

import argparse
import collections
import glob
import json
import os
import random
import re

from ...core import config
from ...sources import uspto_assays as A

# Metric words and units to strip when guessing the biological target from an
# assay column header. Derived, never authoritative — see the README sheet.
_METRIC = re.compile(
    r"\b(ic\s*50|ec\s*50|ed\s*50|gi\s*50|cc\s*50|ki|kd|kb|pic50|pec50|"
    r"binding|inhibition|activity|potency|affinity|assay|value|data|"
    r"flux|fret|spa|htrf|lance|elisa|cell(?:ular)?|enzym(?:e|atic)|"
    r"human|mean|avg|average)\b", re.I)
# Word-bounded. Without the boundaries and with re.I, `uM` matched inside
# "Human" and "Human Whole Blood LTB4" came out as "H an Whole Blood LTB4".
_UNITY = re.compile(r"[(\[]?\s*\b(?:nM|µM|μM|uM|mM|pM)\b\s*[)\]]?|%", re.I)


def target_from(assay_name: str) -> str:
    """Best guess at the biological target named in an assay column header."""
    s = _UNITY.sub(" ", assay_name or "")
    s = _METRIC.sub(" ", s)
    s = re.sub(r"[,;:*]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip(" -–—()[]")
    return s


def slug(assay_name: str, unit: str | None) -> str:
    s = re.sub(r"[^0-9A-Za-z]+", "_", (assay_name or "assay")).strip("_").lower()
    s = re.sub(r"_+", "_", s)[:48] or "assay"
    return f"{s}_{(unit or 'unitless').replace('%', 'pct')}"


def load_structures(pid: str) -> dict[str, dict]:
    p = config.OUTPUT_DIR / "text_extraction" / pid / "example_index.json"
    if not p.exists():
        return {}
    try:
        d = json.loads(p.read_text())
    except (json.JSONDecodeError, OSError):
        return {}
    recs = d if isinstance(d, list) else list(d.values())
    out: dict[str, dict] = {}
    for r in recs:
        if not isinstance(r, dict):
            continue
        cid = A.normalize_cid(str(r.get("compound_id") or "")).upper()
        if cid and cid not in out:
            out[cid] = r
    return out


def collect() -> tuple[list[dict], list[dict]]:
    """(long rows, compound rows) for every patent with structures AND assays."""
    long_rows: list[dict] = []
    by_compound: dict[tuple[str, str], dict] = {}
    for p in sorted(glob.glob(str(config.OUTPUT_DIR / "text_extraction" / "*"
                                  / "example_index.json"))):
        pid = os.path.basename(os.path.dirname(p))
        xml = config.OUTPUT_DIR / "uspto_xml" / f"{pid}.xml"
        if not xml.exists():
            continue
        struct = load_structures(pid)
        if not struct:
            continue
        for rec in A.extract_from_patent(xml.read_text(errors="ignore")):
            cid = A.normalize_cid(rec.cid or "").upper()
            s = struct.get(cid)
            if not s or not rec.is_usable:
                continue
            key = (pid, cid)
            if key not in by_compound:
                by_compound[key] = {
                    "patent_id": pid,
                    "compound_id": rec.cid,
                    "iupac_name": s.get("iupac_name") or "",
                    "canonical_smiles": s.get("canonical_smiles") or "",
                    "inchikey": s.get("inchikey") or "",
                    "iupac_source": s.get("iupac_source") or "",
                    "structure_method": s.get("extraction_method") or "",
                    "_assays": {},
                }
            row = by_compound[key]
            long_rows.append({
                "patent_id": pid,
                "compound_id": rec.cid,
                "iupac_name": row["iupac_name"],
                "canonical_smiles": row["canonical_smiles"],
                "inchikey": row["inchikey"],
                "target": target_from(rec.assay_name),
                "assay": rec.assay_name,
                "value": rec.value_numeric,
                "unit": rec.unit or "",
                "qualifier": rec.qualifier or "",
                "n_runs": rec.n_runs,
                "range_low": rec.range_lo,
                "range_high": rec.range_hi,
                "value_kind": ("number" if rec.value_numeric is not None
                               else "range_from_published_bin"),
                "published_as": rec.value_text or "",
                "letter_grade": rec.letter_grade or "",
                "source_table": rec.table_id,
                "column_header": rec.column_header or rec.assay_name,
                "unit_source": rec.unit_source,
                "extraction_source": rec.source,
            })
            row["_assays"].setdefault(slug(rec.assay_name, rec.unit), rec)
    return long_rows, list(by_compound.values())


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--n", type=int, default=3000, help="compounds to sample")
    ap.add_argument("--seed", type=int, default=20260728)
    ap.add_argument("--out", default="docs/reports/patent_compounds_sample.xlsx")
    args = ap.parse_args()

    long_rows, compounds = collect()
    print(f"{len(compounds)} compounds with a structure AND >=1 usable assay "
          f"({len(long_rows)} assay records)")

    random.seed(args.seed)
    sample = compounds if len(compounds) <= args.n else random.sample(compounds, args.n)
    keep = {(r["patent_id"], A.normalize_cid(r["compound_id"]).upper()) for r in sample}
    long_keep = [r for r in long_rows
                 if (r["patent_id"], A.normalize_cid(r["compound_id"]).upper()) in keep]

    # Wide sheet: one named column group per assay actually present in the sample.
    slugs: list[str] = []
    for r in sample:
        for s in r["_assays"]:
            if s not in slugs:
                slugs.append(s)
    slugs.sort()
    base = ["patent_id", "compound_id", "iupac_name", "canonical_smiles", "inchikey",
            "target", "n_assays", "iupac_source", "structure_method"]
    header = list(base)
    for s in slugs:
        header += [s, f"{s}_qualifier", f"{s}_n_runs", f"{s}_range_low",
                   f"{s}_range_high", f"{s}_assay"]

    wide: list[list] = []
    for r in sample:
        targets = [target_from(rec.assay_name) for rec in r["_assays"].values()]
        seen: list[str] = []
        for t in targets:
            if t and t not in seen:
                seen.append(t)
        row = {c: "" for c in header}
        row.update({k: v for k, v in r.items() if k in base})
        row["target"] = "; ".join(seen)
        row["n_assays"] = len(r["_assays"])
        for s, rec in r["_assays"].items():
            row[s] = rec.value_numeric if rec.value_numeric is not None else ""
            row[f"{s}_qualifier"] = rec.qualifier or ""
            row[f"{s}_n_runs"] = rec.n_runs if rec.n_runs is not None else ""
            row[f"{s}_range_low"] = rec.range_lo if rec.range_lo is not None else ""
            row[f"{s}_range_high"] = rec.range_hi if rec.range_hi is not None else ""
            row[f"{s}_assay"] = rec.assay_name
        wide.append([row[c] for c in header])

    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="1F6F6B")

    def sheet(name, cols, rows, widths):
        ws = wb.create_sheet(name)
        ws.append(list(cols))
        for c in range(1, len(cols) + 1):
            cell = ws.cell(row=1, column=c)
            cell.font, cell.fill = hdr_font, hdr_fill
            cell.alignment = Alignment(vertical="center", wrap_text=False)
        for r in rows:
            ws.append(list(r))
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "C2"
        ws.auto_filter.ref = ws.dimensions
        return ws

    sheet("compounds", header, wide,
          [14, 13, 62, 52, 30, 26, 9, 14, 18] + [15] * (len(header) - len(base)))

    lcols = ["patent_id", "compound_id", "iupac_name", "canonical_smiles", "inchikey",
             "target", "assay", "value", "unit", "qualifier", "n_runs",
             "range_low", "range_high", "value_kind", "published_as", "letter_grade",
             "source_table", "column_header", "unit_source", "extraction_source"]
    sheet("assay_long", lcols, [[r.get(c, "") for c in lcols] for r in long_keep],
          [14, 13, 62, 52, 30, 24, 34, 12, 8, 10, 8, 11, 11, 24, 14, 12, 18, 34, 12, 22])

    n_num = sum(1 for r in long_keep if r["value_kind"] == "number")
    n_bin = len(long_keep) - n_num
    pats = sorted({r["patent_id"] for r in sample})
    readme = [
        ["PatentMoleculeDB — compound + assay sample", ""],
        ["", ""],
        ["Rows (compounds sheet)", len(sample)],
        ["Assay records (assay_long sheet)", len(long_keep)],
        ["  ...a number printed in the patent", n_num],
        ["  ...a range, from a grade the patent published", n_bin],
        ["Patents represented", len(pats)],
        ["Sampled from", f"{len(compounds)} compounds having BOTH a structure and >=1 assay"],
        ["Random seed", args.seed],
        ["", ""],
        ["COLUMN", "MEANING"],
        ["patent_id", "The patent this compound and its values are attributed to."],
        ["compound_id", "The identifier the patent uses (its Example number or code)."],
        ["iupac_name", "Name as printed in the patent, after the IUPAC cascade."],
        ["canonical_smiles", "RDKit canonical SMILES derived from the name."],
        ["inchikey", "InChIKey derived from the structure."],
        ["target", "DERIVED from the assay column header by stripping metric and "
                   "unit words. A convenience for filtering, not curated - trust "
                   "`assay` / `column_header` over this."],
        ["n_assays", "How many distinct assays this compound has in this patent."],
        ["<assay>_<unit>", "The measured value, in the unit named in the column."],
        ["<assay>_qualifier", "'<', '>' etc. where the patent qualified the value."],
        ["<assay>_n_runs", "Replicate count, where the patent printed one."],
        ["<assay>_range_low/high", "Set INSTEAD of a value when the patent published "
                                   "a grade ('++', 'B') rather than a number. The "
                                   "interval is that table's own published key."],
        ["<assay>_assay", "The assay description as the patent printed it."],
        ["value_kind (long)", "'number' = printed in the patent. "
                              "'range_from_published_bin' = the patent printed a "
                              "grade; we resolved it against that table's key."],
        ["published_as (long)", "The raw cell text, so any value can be traced back."],
        ["source_table (long)", "The patent's own table id the value came from."],
        ["", ""],
        ["READ THIS BEFORE DOCKING ANYTHING", ""],
        ["Assay values are exact. Structures are not.",
         "Benchmarked against your hand-curated US8952177 CSV (190 compounds): "
         "every one of 359 measurements matches exactly - value, qualifier AND "
         "replicate count, 190/190 FLAP Ki and 169/169 HWB LTB4. The IUPAC names "
         "on the same rows are 88% right: 168/190 are the same structure, and 20 "
         "(10.5%) are a DIFFERENT compound from the patent pasted onto the wrong "
         "compound_id."],
        ["What that means for a row here",
         "A wrong row is not a wrong number - it is a correct number attached to "
         "the wrong molecule, which docking cannot detect. Treat compound_id + "
         "patent_id + the assay value as reliable, and iupac_name/SMILES as ~90% "
         "reliable pending a fix to compound-id alignment. Spot-check before "
         "committing compute to any single structure."],
        ["", ""],
        ["WHAT THESE NUMBERS DO NOT MEAN", ""],
        ["Ranges are not point values",
         "Where a patent published a grade there is no number in the document. The "
         "interval is real and usable for triage; it will never sharpen, because "
         "only the assay can produce the number."],
        ["Enantiomers are separate rows",
         "(R) and (S) forms are kept as distinct entries, per your Apr 9 answer."],
        ["Markush is NOT enumerated",
         "Every row here is a compound the patent names explicitly. Compounds that "
         "exist only inside a generic structure are absent, and that is the largest "
         "known gap in coverage."],
        ["Structure coverage limits the sample",
         "Assay extraction covers 73 patents; structures have been generated for 20. "
         "The sample is drawn from the intersection, not from all assay data."],
    ]
    sheet("README", ["FIELD", "DETAIL"], readme, [42, 96])
    ws = wb["README"]
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = None
    for row in ws.iter_rows(min_row=2):
        row[1].alignment = Alignment(wrap_text=True, vertical="top")

    del wb["Sheet"]
    out = config.REPO_ROOT / args.out
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"wrote {out}")
    print(f"  compounds : {len(sample)} rows x {len(header)} cols "
          f"({len(slugs)} distinct assays)")
    print(f"  assay_long: {len(long_keep)} rows ({n_num} numbers, {n_bin} ranges)")
    print(f"  patents   : {len(pats)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
