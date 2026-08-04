"""A workbook for EYEBALLING what the text pull produced, patent by patent.

    python3 -m patentdb.scripts.eval.sample_workbook                    # 5 at random
    python3 -m patentdb.scripts.eval.sample_workbook -n 8 --seed 7
    python3 -m patentdb.scripts.eval.sample_workbook --patents US8952177,US9302989
    python3 -m patentdb.scripts.eval.sample_workbook --any             # ignore structures

Four sheets, because the interesting question is not "how many rows" but
"does a row mean what it says":

  Summary    one line per patent — records, compounds, the join rate
  Compounds  one line per (patent, compound) — the JOIN, side by side:
             the id the assay table used, the name we found for it, the
             structure that name resolved to, and what BindingDB says
  Records    every assay measurement, with the raw cell text beside the
             parsed number so a misparse is visible without opening the XML
  Unjoined   compounds carrying assay values that reached NO structure

SAMPLED FROM PATENTS WHERE THE STRUCTURE RESOLVER HAS RUN, unless `--any`.
Only 20 of 123 patents have an `example_index.json`; the rest have assays and
no molecules, so a random draw across all of them would show mostly blank
name/SMILES columns and read as broken extraction when it is really work that
was never run. `--any` gives the true random draw when that is what you want.
"""
from __future__ import annotations

import argparse
import json
import logging
import os
import random
import sys

logger = logging.getLogger(__name__)

_HDR = {"bold": True}


def _structures(pid: str) -> dict:
    from ...core import config
    p = config.OUTPUT_DIR / "text_extraction" / pid / "example_index.json"
    if not p.exists():
        return {}
    try:
        d = json.loads(p.read_text())
    except (OSError, ValueError):
        return {}
    return d if isinstance(d, dict) else {}


def _norm(cid: str) -> str:
    s = str(cid or "").strip()
    return s.lstrip("0").upper() or s.upper()


def _records(pid: str, xml: str) -> list:
    """Both reading tiers, deduplicated — the same union `pipeline_bench` uses."""
    os.environ.setdefault("REPAIR", "0")
    from ...repair.loop import repair_patent
    from ...sources.uspto_assays import extract_from_patent

    out, seen = [], set()
    for fn in (lambda: extract_from_patent(xml), lambda: repair_patent(pid, xml)[0]):
        try:
            rs = list(fn())
        except Exception as e:
            logger.warning("%s: a tier raised %r", pid, e)
            continue
        for r in rs:
            if not getattr(r, "is_usable", False):
                continue
            k = (r.cid, r.assay_name, r.value_numeric, r.unit, r.table_id)
            if k not in seen:
                seen.add(k)
                out.append(r)
    return out


def build(pids: list[str], dest) -> dict:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    from ...core import config
    try:
        from .reference_bench import load_bindingdb
        bdb, _ = load_bindingdb()
    except Exception as e:
        logger.warning("BindingDB unavailable (%r)", e)
        bdb = {}

    xml_dir = config.OUTPUT_DIR / "uspto_xml"
    wb = Workbook()
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_cmp = wb.create_sheet("Compounds")
    ws_rec = wb.create_sheet("Records")
    ws_un = wb.create_sheet("Unjoined")

    ws_sum.append(["patent", "assay_records", "compounds", "with_name",
                   "with_structure", "join_%", "distinct_assays",
                   "bdb_ligands", "bdb_matched"])
    ws_cmp.append(["patent", "compound_id", "n_records", "assays", "has_name",
                   "iupac_name", "canonical_smiles", "inchikey",
                   "in_bindingdb", "iupac_source", "extraction_method"])
    ws_rec.append(["patent", "compound_id", "assay_name", "value_numeric",
                   "unit", "qualifier", "n_runs", "range_lo", "range_hi",
                   "letter_grade", "raw_cell_text", "column_header", "table_id"])
    ws_un.append(["patent", "compound_id", "n_records", "example_assay",
                  "example_value", "unit", "why_unjoined"])

    totals = {"records": 0, "compounds": 0, "named": 0, "structured": 0}
    for pid in pids:
        f = xml_dir / f"{pid}.xml"
        if not f.exists():
            logger.warning("%s: no cached XML", pid)
            continue
        recs = _records(pid, f.read_text(errors="ignore"))
        struct = _structures(pid)
        by_norm = {_norm(k): v for k, v in struct.items()}
        by_cid: dict[str, list] = {}
        for r in recs:
            by_cid.setdefault(str(r.cid), []).append(r)
            ws_rec.append([pid, r.cid, r.assay_name, r.value_numeric, r.unit,
                           getattr(r, "qualifier", ""), getattr(r, "n_runs", ""),
                           getattr(r, "range_lo", ""), getattr(r, "range_hi", ""),
                           getattr(r, "letter_grade", ""),
                           getattr(r, "value_text", ""),
                           getattr(r, "column_header", ""), r.table_id])

        iks = set(bdb.get(pid, ()))
        named = structured = matched = 0
        for cid, rs in sorted(by_cid.items()):
            info = struct.get(cid) or by_norm.get(_norm(cid)) or {}
            name = (info.get("iupac_name") or "").strip()
            ik = (info.get("inchikey") or "").strip()
            named += bool(name)
            structured += bool(ik)
            in_b = ""
            if ik and iks:
                in_b = "yes" if ik in iks else "no"
                matched += ik in iks
            ws_cmp.append([pid, cid, len(rs),
                           "; ".join(sorted({r.assay_name for r in rs}))[:200],
                           "yes" if name else "NO", name,
                           info.get("canonical_smiles", ""), ik, in_b,
                           info.get("iupac_source", ""),
                           info.get("extraction_method", "")])
            if not ik:
                r0 = rs[0]
                ws_un.append([pid, cid, len(rs), r0.assay_name, r0.value_numeric,
                              r0.unit,
                              "no name found for this id" if not name
                              else "name found but would not resolve"])

        n = len(by_cid) or 1
        ws_sum.append([pid, len(recs), len(by_cid), named, structured,
                       round(100.0 * structured / n, 1),
                       len({r.assay_name for r in recs}),
                       len(iks), matched])
        totals["records"] += len(recs)
        totals["compounds"] += len(by_cid)
        totals["named"] += named
        totals["structured"] += structured

    n = totals["compounds"] or 1
    ws_sum.append([])
    ws_sum.append(["TOTAL", totals["records"], totals["compounds"],
                   totals["named"], totals["structured"],
                   round(100.0 * totals["structured"] / n, 1)])

    for ws, widths in ((ws_sum, [16, 14, 11, 10, 14, 8, 15, 12, 12]),
                       (ws_cmp, [14, 14, 10, 46, 9, 60, 52, 30, 12, 18, 26]),
                       (ws_rec, [14, 14, 40, 14, 8, 9, 8, 10, 10, 12, 24, 34, 18]),
                       (ws_un, [14, 14, 10, 34, 13, 8, 34])):
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        for c in ws[1]:
            c.font = Font(**_HDR)
            c.alignment = Alignment(vertical="top")
        ws.freeze_panes = "A2"
        if ws.max_row > 1:
            ws.auto_filter.ref = ws.dimensions

    dest.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dest)
    return totals


def main() -> int:
    logging.basicConfig(level=logging.WARNING, format="%(message)s")
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--patents", help="comma-separated; overrides sampling")
    ap.add_argument("-n", type=int, default=5, help="how many to sample")
    ap.add_argument("--seed", type=int, default=None)
    ap.add_argument("--any", action="store_true",
                    help="sample from ALL patents, not only those with structures")
    ap.add_argument("--out", default="docs/reports/sample_extraction.xlsx")
    a = ap.parse_args()
    from ...core import config

    xml_dir = config.OUTPUT_DIR / "uspto_xml"
    every = sorted(p.stem for p in xml_dir.glob("*.xml"))
    if a.patents:
        pids = [p.strip().upper() for p in a.patents.split(",")]
    else:
        pool = every if a.any else [p for p in every if _structures(p)]
        if not pool:
            print("no patents with a resolved example_index; pass --any", file=sys.stderr)
            return 1
        rng = random.Random(a.seed)
        pids = sorted(rng.sample(pool, min(a.n, len(pool))))
    print(f"sampling {len(pids)}: {', '.join(pids)}", file=sys.stderr)

    dest = config.REPO_ROOT / a.out
    t = build(pids, dest)
    n = t["compounds"] or 1
    print(f"\nrecords    : {t['records']}")
    print(f"compounds  : {t['compounds']}")
    print(f"with name  : {t['named']}")
    print(f"structured : {t['structured']}  ({100.0*t['structured']/n:.1f}% joined)")
    print(f"\nwrote {dest}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
