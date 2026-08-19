"""Turn the current reader dump into a workbook for eyeballing.

    python3 -m patentdb3.to_excel            # whatever the last dump run produced
    python3 -m patentdb3.to_excel --no-bdb   # skip the BindingDB sheet

READS THE DUMP PATH FROM THE MANIFEST, NEVER FROM A HARDCODED STRING.
`verify.py --dump` rewrites `out/latest.json` on every run; this reads that and
follows it. So the workbook is always built from the newest extraction, and if
the dump ever moves there is exactly one place that changes. Hardcoding a second
path here is how v2 ended up with consumers quietly reading stale files.

Output goes to ONE place, overwritten: `out/reader_dump.xlsx`.

TOLERANCE. Assay agreement is judged at 1%, not 5%. The reader does not round —
it reports the cell as printed — so the only legitimate gap between our value
and a reference is the reference's own rounding to two or three significant
figures. 5% is wide enough to call a genuinely different measurement a match,
which is the failure mode that lets a scorer improve while the data gets worse.

BindingDB is a REFERENCE. Nothing here writes a BDB value into an extracted
record, and a disagreement is reported, never corrected.
"""
from __future__ import annotations

import collections
import csv
import json
import re
import sys
from pathlib import Path

from .core import config
# One shared table — see `sources.uspto_assays.TO_NM` for why.
from .sources.uspto_assays import TO_NM as _TO_NM


MANIFEST_PATH = config.MANIFEST
XLSX_PATH = config.XLSX
BDB_TSV = config.REPO_ROOT / "output" / "bindingdb" / "our_patents.tsv"

# See the module docstring. We do not round, so only the reference's rounding
# is forgivable.
TOLERANCE = 0.01


def _load() -> tuple[list[dict], dict]:
    if not MANIFEST_PATH.exists():
        sys.exit(f"no manifest at {MANIFEST_PATH} — run `python3 -m patentdb3.verify <PID> --dump` first")
    man = json.loads(MANIFEST_PATH.read_text())
    # The manifest records an ABSOLUTE path, which is right as a record and
    # useless as a lookup once the checkout moves. Same precedence as
    # `images.emit`: the recorded path wins when it exists, and the configured
    # location is reached for only when it has gone away.
    dump = Path(man["dump"])
    if not dump.exists() and Path(config.DUMP).exists():
        dump = Path(config.DUMP)
    if not dump.exists():
        sys.exit(f"manifest names {dump}, which does not exist")
    with dump.open() as fh:
        rows = list(csv.DictReader(fh, delimiter="\t"))
    return rows, man


# What one printed value is worth in nM, the unit BindingDB publishes in.
# Anything not here is not comparable and is left out rather than guessed at —
# a pIC50 or a bare percentage is not a concentration.
def _to_nM(value, unit):
    """`value` in `unit` as nanomolar, or None when it is not a concentration."""
    f = _TO_NM.get((unit or "").strip())
    if f is None or not value:
        return None
    try:
        return float(value) * f
    except (TypeError, ValueError):
        return None


def _bdb(pids: set[str]) -> dict[tuple[str, str], set[float]]:
    """Reference points, keyed TWO ways: by compound number and by structure.

    THE COMPOUND NUMBER IS BINDINGDB'S, NOT THE PATENT'S. It is parsed out of a
    free-text ligand name, and the two numbering schemes are not the same
    document's: US8722692 Example 7 — `JFPWYXWNCPJSCN-UHFFFAOYSA-N` — appears in
    BindingDB as `US8722692, 576`. Joining on it therefore compares the wrong
    compound, quietly, because a wrong pairing still produces a number.

    Measured over the corpus, on the 11,902 compounds reachable by either key:

        values agree on the id join        4,003   33.6%
        values agree on the structure join 11,308  95.0%
        reachable ONLY by structure        7,892

    So the number this sheet reports is the structure join, and the id join is
    kept beside it as the thing it replaced. An InChIKey cannot drift: two rows
    carrying the same one are the same molecule.

    The ligand name holds `PATENT, CID` pairs joined by `::`, and one row is
    routinely cited against several patents at once — 97% of US8952177's rows
    are that shape. So the cid must be read from THIS patent's own pair; taking
    the first number in the string picks up a sibling patent's compound, whose
    numbering is unrelated.
    """
    out: dict[tuple[str, str], set[float]] = collections.defaultdict(set)
    by_structure: dict[tuple[str, str], set[float]] = collections.defaultdict(set)
    if not BDB_TSV.exists():
        return out, by_structure
    csv.field_size_limit(10 ** 9)
    with BDB_TSV.open(newline="") as fh:
        rd = csv.reader(fh, delimiter="\t", quoting=csv.QUOTE_NONE)
        hdr = next(rd)
        iP, iN = hdr.index("Patent Number"), hdr.index("BindingDB Ligand Name")
        iK = hdr.index("Ligand InChI Key")
        vcols = [i for i, h in enumerate(hdr) if h in ("IC50 (nM)", "Ki (nM)")]
        for row in rd:
            if len(row) <= max(iP, iN):
                continue
            pid = next((p for p in pids if p in (row[iP] or "")), None)
            if not pid:
                continue
            m = re.search(rf"{pid},\s*(\d{{1,4}})\b", row[iN] or "")
            if not m:
                continue
            for i in vcols:
                v = (row[i] or "").strip().lstrip("<>=~ ")
                try:
                    out[(pid, m.group(1))].add(float(v))
                except ValueError:
                    pass
            # THE SAME ROW, KEYED ON CHEMISTRY. See `_bdb`'s docstring.
            ik = (row[iK] or "").strip()
            if ik:
                for i in vcols:
                    v = (row[i] or "").strip().lstrip("<>=~ ")
                    try:
                        by_structure[(pid, ik)].add(float(v))
                    except ValueError:
                        pass
    return out, by_structure


def main(argv: list[str]) -> int:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    rows, man = _load()
    pids = {r["patent_id"] for r in rows}
    wb = Workbook()

    ws = wb.active
    ws.title = "records"
    cols = list(rows[0].keys()) if rows else []
    ws.append(cols)
    for r in rows:
        ws.append([r[c] for c in cols])
    for c in ws[1]:
        c.font = Font(bold=True)
    ws.freeze_panes = "A2"
    for i, c in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = max(11, min(38, len(c) + 22))

    # one row per assay: how many measurements, what units, which tables
    sm = wb.create_sheet("assays")
    sm.append(["patent_id", "assay_name", "records", "units", "tables"])
    agg = collections.defaultdict(lambda: [0, set(), set()])
    for r in rows:
        a = agg[(r["patent_id"], r["assay_name"])]
        a[0] += 1
        a[1].add(r["unit"])
        a[2].add(r["table_id"])
    for (p, a), (n, u, t) in sorted(agg.items()):
        sm.append([p, a, n, ", ".join(sorted(x for x in u if x)), ", ".join(sorted(x for x in t if x))])
    for c in sm[1]:
        c.font = Font(bold=True)
    sm.freeze_panes = "A2"
    sm.column_dimensions["B"].width = 42

    _joined_sheet(wb, rows, man)
    _trail_sheets(wb, man)

    if "--no-bdb" not in argv:
        ref, ref_by_structure = _bdb(pids)
        ours = collections.defaultdict(set)
        # EVERY UNIT WE CAN CONVERT, NOT JUST uM. This read
        # `r["unit"] == "uM"` and BindingDB publishes in nM — so the filter
        # threw away the exact column BDB cites (41,959 nM rows against 40,893
        # uM) and compared whatever micromolar column the patent also printed.
        # US10004738 scored 0 of 49 because BDB cites `BACE1 Ki (nM)` — cid 1
        # prints 83.0 and BDB says 83 — while this compared `at 10 uM (%)`, a
        # percent-inhibition column. Agreement on the overlap read 67.5%; with
        # nM included and nothing else changed it is 96.2%.
        #
        # The docstring said this judged assay agreement. It judged one unit.
        for r in rows:
            nm = _to_nM(r.get("value_numeric"), r.get("unit"))
            if nm is not None:
                ours[(r["patent_id"], r["cid"])].add(nm)
        # OUR VALUES KEYED ON STRUCTURE TOO, from `structures.tsv`, so the
        # comparison can join on chemistry rather than on a compound number
        # BindingDB assigns itself. 89.9% of its rows for our patents carry an
        # InChIKey and only 16.1% yield a parseable compound number, so the id
        # key was discarding five sixths of the available reference before it
        # compared anything.
        ours_by_structure: dict = collections.defaultdict(set)
        src = Path(man.get("structures") or "")
        if src.exists():
            key_of = {(r["patent_id"], r["cid"]): r["inchikey"]
                      for r in csv.DictReader(src.open(), delimiter="\t")
                      if r.get("inchikey") and r.get("cid")}
            for (p, c), vals in ours.items():
                ik = key_of.get((p, c))
                if ik:
                    ours_by_structure[(p, ik)] |= vals

        sb = wb.create_sheet("bindingdb")
        sb.append(["patent_id", "join", "key", "bdb_nM", "ours_nM",
                   "agrees_1pct", "note"])
        hit = tot = held = 0
        for (p, ik), vals in sorted(ref_by_structure.items()):
            mine = ours_by_structure.get((p, ik), set())
            if not mine:
                continue                      # a compound we have not resolved
            for v in sorted(vals):
                ok = any(abs(v - o) <= TOLERANCE * max(v, 1e-12) for o in mine)
                tot += 1
                hit += ok
                held += 1
                sb.append([p, "structure", ik, v,
                           ", ".join(f"{o:g}" for o in sorted(mine)) or "—",
                           "yes" if ok else "no", ""])
        # The id join is kept beside it, as the thing it replaced.
        id_hit = id_tot = 0
        for (p, c), vals in sorted(ref.items()):
            mine = ours.get((p, c), set())
            for v in sorted(vals):
                ok = any(abs(v - o) <= TOLERANCE * max(v, 1e-12) for o in mine)
                id_tot += 1
                id_hit += ok
                sb.append([p, "compound no.", c, v,
                           ", ".join(f"{o:g}" for o in sorted(mine)) or "—",
                           "yes" if ok else "no",
                           "" if mine else "cid not extracted"])
        for c in sb[1]:
            c.font = Font(bold=True)
        sb.freeze_panes = "A2"
        # TWO NUMBERS, BECAUSE THEY MEASURE DIFFERENT THINGS. A reference value
        # for a compound we never extracted is a COVERAGE miss; counting it as
        # a disagreement reports the reader as wrong for a drawing that has not
        # been through recognition yet. One line printed `8725/10350 = 84.3%`
        # while agreement on the compounds we actually hold was 8725/9072 =
        # 96.2%, and the two were read as the same quantity.
        print(f"bindingdb (structure join): {hit}/{held} agree within "
              f"{TOLERANCE:.0%}" + (f"  ({hit/held:.1%})" if held else ""))
        print(f"  the compound-number join it replaces: {id_hit}/{id_tot}"
              + (f"  ({id_hit/id_tot:.1%})" if id_tot else "")
              + " — BindingDB numbers compounds itself, so that key compares "
                "the wrong molecule")

    XLSX_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(XLSX_PATH)
    print(f"{len(rows):,} records from {man['written_at']} ({', '.join(sorted(pids))})")
    print(f"-> {XLSX_PATH}")
    return 0


def _range_text(lo, hi) -> str:
    """A bin's interval, written the way the patent states it.

    `3-7`, `<7`, `>3`. An open side is a real statement — a compound the
    document places under 3 nM is bounded above and nowhere below — so it is
    written as the inequality it is rather than padded with a fake endpoint.
    """
    def num(v):
        try:
            return f"{float(v):g}"
        except (TypeError, ValueError):
            return ""
    a, b = num(lo), num(hi)
    if a and b:
        return f"{a}-{b}"
    if b:
        return f"<{b}"
    if a:
        return f">{a}"
    return ""


def _joined_sheet(wb, dump_rows, man) -> None:
    """THE DELIVERABLE. One row per compound: identity joined to its assays.

    Everything else in this workbook is working material. This sheet is the
    thing the pipeline exists to produce, and the only one that answers "what
    did you get out of this patent" without the reader assembling it.

    ASSAYS ARE TWO ALIGNED LISTS, NOT A COLUMN EACH. A column per assay needs
    505 of them — that is how many distinct assay names this corpus holds —
    so the first version showed the commonest 40 and silently hid the rest.
    It also spent 160 columns on a compound that has, at most, 12 values;
    measured, 13,827 compounds have exactly one and only 217 have eight.

    So the assay names go in one cell and their values in the next, in the
    same order. Every assay a compound has is visible, nothing is truncated,
    and reading across one row tells you what was measured and what it came
    to. Sorting or filtering still works on the columns that matter — patent,
    route, verdict.
    """
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    src = Path(man.get("structures") or "")
    if not src.exists():
        src = Path(config.STRUCTURES)
    if not src.exists():
        return
    csv.field_size_limit(10 ** 8)
    with src.open(newline="") as fh:
        st = list(csv.DictReader(fh, delimiter="\t"))

    def rank(r):
        return (r.get("mass_check") == "agrees", bool(r.get("inchikey")))
    ident: dict = {}
    for r in st:
        k = (r["patent_id"], r.get("cid") or "")
        if not k[1]:
            continue
        if k not in ident or rank(r) > rank(ident[k]):
            ident[k] = r

    # {compound -> {assay -> [value, ...]}}, insertion order kept so the two
    # lists in the sheet line up exactly.
    meas: dict = collections.defaultdict(dict)
    for r in dump_rows:
        k = (r["patent_id"], r.get("cid") or "")
        a = (r.get("assay_name") or "").strip()
        if not k[1] or not a:
            continue
        # DO NOT REPEAT A UNIT THE ASSAY NAME ALREADY CARRIES. Patents
        # routinely head a column `BACE1 Ki (nM)`, and appending the parsed
        # unit again gave `BACE1 Ki (nM) nM`.
        unit = (r.get("unit") or "").strip()
        label = a if (not unit or unit.lower() in a.lower()) else f"{a} {unit}"
        # A GRADE IS SHOWN AS THE RANGE IT MEANS, NOT AS THE LETTER.
        #
        # The letter was taken before the range, so the range was never
        # reached and this sheet — the deliverable — printed `B, B, B` for a
        # compound whose three B's the patent defines as 3-7 nM, 0.5-5 uM and
        # 15-25 uM. 10,238 of 45,676 rows showed a bare symbol, which is the
        # whole of the bin work invisible exactly where it is read. The
        # `records` sheet still carries `letter_grade` for provenance.
        val = (r.get("value_numeric") or "").strip()
        if not val:
            val = _range_text(r.get("range_lo"), r.get("range_hi"))
        if not val:
            val = (r.get("letter_grade") or "").strip()
        if not val:
            continue
        q = (r.get("qualifier") or "").strip()
        n = (r.get("n_runs") or "").strip()
        cell = (q + val) if q else val
        if n:
            cell += f" (n={n})"
        meas[k].setdefault(label, []).append(cell)

    head = ["patent_id", "cid", "route", "name_iupac", "smiles", "inchikey",
            "verdict", "n_assays", "assays", "values"]
    ws = wb.create_sheet("compounds")
    ws.append(head)
    for k in sorted(set(ident) | set(meas)):
        r = ident.get(k, {})
        v = ("confirmed" if r.get("mass_check") == "agrees" else
             "DISPUTED" if r.get("mass_check") == "contradicts" else
             "resolved" if r.get("inchikey") else
             "markush" if r.get("markush") == "True" else "unresolved")
        route = {"table": "name in a table cell",
                 "cid_first": "search from the compound id"}.get(
                     r.get("source", ""), r.get("source", "") or "\u2014")
        if r.get("markush") == "True":
            route = "markush assembly"
        m = meas.get(k, {})
        ws.append([k[0], k[1], route, (r.get("name") or "")[:300],
                   (r.get("smiles") or "")[:400], r.get("inchikey", ""), v,
                   len(m),
                   ", ".join(m) or "",
                   ", ".join("; ".join(vals) for vals in m.values()) or ""])
    for c in ws[1]:
        c.font = Font(bold=True)
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = ws.dimensions
    widths = {"patent_id": 15, "cid": 10, "route": 27, "name_iupac": 52,
              "smiles": 52, "inchikey": 29, "verdict": 13, "n_assays": 10,
              "assays": 60, "values": 34}
    for i, c in enumerate(head, 1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(c, 14)
    # TEXT, NOT NUMBERS. A compound with one assay yields a values cell like
    # `83.0`, which Excel silently retypes as a number and then renders as
    # `####` whenever the column is narrower than its own idea of the format.
    # These are lists that happen to have one element; they are never numbers.
    for row in ws.iter_rows(min_row=2, min_col=9, max_col=10):
        for cell in row:
            cell.number_format = "@"
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    print(f"compounds: {ws.max_row - 1:,} rows, assays as aligned lists")


def _trail_sheets(wb, man) -> None:
    """One sheet per identity route: compound id in, structure and evidence out.

    THE POINT IS TO BE CHECKABLE BY HAND. `structures.tsv` already holds the
    whole trail in 28 columns, but 38,871 rows of it in one file answers no
    question. Split by the route that produced the row, a chemist can take a
    compound number, see which route claimed it, what text or picture it was
    claimed from, and whether the patent's own mass agrees — without reading
    any code.

    `verdict` is the column to sort on. It is the only one that is not our
    own opinion: it is what the patent printed.
    """
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    src = Path(man.get("structures") or "")
    if not src.exists():
        src = Path(config.STRUCTURES)
    if not src.exists():
        print("no structures.tsv — trail sheets skipped")
        return

    csv.field_size_limit(10 ** 8)
    with src.open(newline="") as fh:
        rows = list(csv.DictReader(fh, delimiter="\t"))

    # Only two sources exist in the dump — `table` and `cid_first` — so the
    # split that answers a question is by WHAT HAPPENED, not by which module
    # ran. `unresolved` is deliberately its own sheet: it is the work queue.
    ROUTES = {
        "trail_confirmed":  lambda r: r.get("mass_check") == "agrees",
        "trail_disputed":   lambda r: r.get("mass_check") == "contradicts",
        "trail_markush":    lambda r: r.get("markush") == "True",
        "trail_resolved":   lambda r: bool(r.get("inchikey")),
        "trail_unresolved": lambda r: True,
    }
    COLS = ["patent_id", "cid", "verdict", "inchikey", "smiles", "name",
            "source", "reason", "mass_check", "mass_delta",
            "markush_kind", "markush_reason", "markush_parts",
            "drawn_ref", "drawn_file", "repair", "heading_transform",
            "dewrap", "raw_cell", "table_id", "row_index", "column_signal"]

    def _verdict(r):
        if r.get("mass_check") == "agrees":
            return "confirmed by printed mass"
        if r.get("mass_check") == "contradicts":
            return "CONTRADICTED by printed mass"
        if r.get("inchikey"):
            return "resolved, nothing to check against"
        if r.get("markush") == "True":
            return "markush — no single structure"
        if r.get("drawn_ref"):
            return "drawn only — needs image recognition"
        return "not resolved"

    used = set()
    for title, keep in ROUTES.items():
        sel = [r for r in rows if keep(r) and id(r) not in used]
        used.update(id(r) for r in sel)
        ws = wb.create_sheet(title)
        ws.append(COLS)
        for r in sel:
            r = {**r, "verdict": _verdict(r)}
            ws.append([(r.get(c) or "")[:400] if isinstance(r.get(c), str)
                       else r.get(c) for c in COLS])
        for c in ws[1]:
            c.font = Font(bold=True)
        ws.freeze_panes = "C2"
        ws.auto_filter.ref = ws.dimensions
        for i, c in enumerate(COLS, 1):
            ws.column_dimensions[get_column_letter(i)].width = (
                46 if c in ("smiles", "name", "raw_cell", "markush_parts")
                else max(11, min(30, len(c) + 6)))
        print(f"{title}: {len(sel):,} rows")

    left = [r for r in rows if id(r) not in used]
    if left:
        print(f"  (+{len(left):,} rows in no route sheet — check ROUTES)")


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
